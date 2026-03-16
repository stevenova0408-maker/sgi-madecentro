# ==========================================================
# IMPORTS PRINCIPALES
# ==========================================================

from fastapi import FastAPI, HTTPException, UploadFile, File, Request, Form, Query
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from starlette.middleware.sessions import SessionMiddleware

from sqlalchemy import (
    create_engine,
    Column,
    Integer,
    String,
    Boolean,
    DateTime,
    ForeignKey,
    Text,
    func,
    UniqueConstraint
)
from sqlalchemy.orm import declarative_base, sessionmaker, relationship

from pydantic import BaseModel
from datetime import datetime, timedelta

import pandas as pd
import os
import re
import logging
import shutil
import smtplib
import uuid  # 🔥 NECESARIO para token_validacion
from email.message import EmailMessage

# ReportLab / PDF / QR (evita errores futuros)
from reportlab.graphics.barcode import qr, eanbc, code128
from reportlab.graphics.shapes import Drawing
from reportlab.graphics import renderPDF
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib import colors, pagesizes
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, inch
from reportlab.lib.colors import HexColor

# ==========================================================
# BASE DECLARATIVA
# ==========================================================

Base = declarative_base()

# ==========================================================
# CONFIGURACIÓN GENERAL
# ==========================================================

DB_CONFIG = {
    "user": "postgres",
    "password": "1234",  # ⚠️ En producción usar variable de entorno
    "host": "127.0.0.1",  # evitar IPv6 (::1)
    "port": "5432"
}

# ==========================================================
# CONEXIÓN BASE DE DATOS
# ==========================================================

import os
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

DATABASE_URL = os.getenv("DATABASE_URL")

if DATABASE_URL:
    # ======================================================
    # PRODUCCIÓN (RENDER)
    # ======================================================

    # Render usa postgres:// y SQLAlchemy necesita postgresql://
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://")

    engine = create_engine(
        DATABASE_URL,
        pool_pre_ping=True,
        pool_recycle=1800
    )

else:
    # ======================================================
    # LOCAL
    # ======================================================

    DATABASE_URL = (
        f"postgresql://{DB_CONFIG['user']}:"
        f"{DB_CONFIG['password']}@"
        f"{DB_CONFIG['host']}:"
        f"{DB_CONFIG['port']}/sgi_madecentro_db"
    )

    engine = create_engine(
        DATABASE_URL,
        pool_pre_ping=True,
        pool_recycle=1800
    )

SessionLocal = sessionmaker(
    autocommit=False,
    autoflush=False,
    bind=engine
)

# ==========================================================
# LOGS
# ==========================================================

import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("SGI")

# ==========================================================
# BASES DISPONIBLES (MULTI-PLANTA)
# ==========================================================

if os.getenv("DATABASE_URL"):

    # ======================================================
    # PRODUCCIÓN (RENDER)
    # ======================================================

    DATABASE_URL = os.getenv("DATABASE_URL").replace("postgres://", "postgresql://")

    DATABASES = {
        "050": DATABASE_URL,
        "051": DATABASE_URL,
        "052": DATABASE_URL,
        "053": DATABASE_URL,
        "064": DATABASE_URL,
        "065": DATABASE_URL,
        "piloto": DATABASE_URL
    }

else:

    # ======================================================
    # LOCAL
    # ======================================================

    DATABASES = {
        "050": f"postgresql+psycopg2://{DB_CONFIG['user']}:{DB_CONFIG['password']}@{DB_CONFIG['host']}:{DB_CONFIG['port']}/sgi_050",
        "051": f"postgresql+psycopg2://{DB_CONFIG['user']}:{DB_CONFIG['password']}@{DB_CONFIG['host']}:{DB_CONFIG['port']}/sgi_051",
        "052": f"postgresql+psycopg2://{DB_CONFIG['user']}:{DB_CONFIG['password']}@{DB_CONFIG['host']}:{DB_CONFIG['port']}/sgi_052",
        "053": f"postgresql+psycopg2://{DB_CONFIG['user']}:{DB_CONFIG['password']}@{DB_CONFIG['host']}:{DB_CONFIG['port']}/sgi_053",
        "064": f"postgresql+psycopg2://{DB_CONFIG['user']}:{DB_CONFIG['password']}@{DB_CONFIG['host']}:{DB_CONFIG['port']}/sgi_064",
        "065": f"postgresql+psycopg2://{DB_CONFIG['user']}:{DB_CONFIG['password']}@{DB_CONFIG['host']}:{DB_CONFIG['port']}/sgi_065",

        # 🔥 BASE PILOTO
        "piloto": f"postgresql+psycopg2://{DB_CONFIG['user']}:{DB_CONFIG['password']}@{DB_CONFIG['host']}:{DB_CONFIG['port']}/sgi_piloto",
    }

# ==========================================================
# CREAR ENGINES (1 POR PLANTA)
# ==========================================================

engines = {
    codigo: create_engine(
        url,
        echo=False,
        pool_pre_ping=True,
        pool_recycle=1800
    )
    for codigo, url in DATABASES.items()
}

# ==========================================================
# CREAR SESSIONMAKERS
# ==========================================================

SessionLocals = {
    codigo: sessionmaker(
        autocommit=False,
        autoflush=False,
        bind=engine
    )
    for codigo, engine in engines.items()
}

# ==========================================================
# FUNCIÓN CENTRAL MULTI-PLANTA
# ==========================================================

from fastapi import HTTPException

def get_db(planta_codigo: str):
    """
    Devuelve la sesión correcta según la planta activa.
    """

    if not planta_codigo:
        raise HTTPException(
            status_code=401,
            detail="No hay planta seleccionada en sesión"
        )

    if planta_codigo not in SessionLocals:
        raise HTTPException(
            status_code=400,
            detail=f"Planta inválida: {planta_codigo}"
        )

    return SessionLocals[planta_codigo]()
# ==========================================================
# CREAR TABLAS EN TODAS LAS PLANTAS (si no existen)
# ==========================================================

for codigo, engine in engines.items():
    Base.metadata.create_all(bind=engine)
    print(f"✔ Tablas verificadas/creadas en planta {codigo}")

# ==========================================================
# SINCRONIZAR ESTRUCTURA EN TODAS LAS PLANTAS
# ==========================================================

from sqlalchemy import text

for codigo, engine in engines.items():
    try:
        with engine.connect() as conn:
            conn.execute(text("""
                ALTER TABLE entrega_cedi
                ADD COLUMN IF NOT EXISTS token_validacion VARCHAR(64);
            """))
            conn.commit()
        print(f"✔ entrega_cedi sincronizada en planta {codigo}")
    except Exception as e:
        print(f"⚠ No se pudo sincronizar entrega_cedi en {codigo}: {e}")

# ==========================================================
# CREAR SESSIONMAKERS
# ==========================================================

SessionLocals = {
    codigo: sessionmaker(
        autocommit=False,
        autoflush=False,
        bind=engine
    )
    for codigo, engine in engines.items()
}

# ==========================================================
# FUNCIÓN CENTRAL MULTI-PLANTA
# ==========================================================

from fastapi import HTTPException

def get_db(planta_codigo: str):
    """
    Devuelve una sesión de base de datos según la planta activa.
    """

    if not planta_codigo:
        raise HTTPException(
            status_code=401,
            detail="No hay planta seleccionada en sesión"
        )

    if planta_codigo not in SessionLocals:
        raise HTTPException(
            status_code=400,
            detail=f"Planta inválida: {planta_codigo}"
        )

    return SessionLocals[planta_codigo]()

# ==========================================================
# CONFIGURACIÓN CORREO SGI
# ==========================================================

SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587

EMAIL_REMITENTE = "entregasmadeservicios@gmail.com"
EMAIL_PASSWORD = "fluk sktp yxur yttt"

# Correos internos que siempre irán en copia
CORREOS_FIJOS = [
    "logistica@empresa.com",
    "controlinterno@empresa.com"
]


# ==========================================================
# APP FASTAPI
# ==========================================================

app = FastAPI()


# ==========================================================
# CONFIGURACIÓN DE SESIÓN
# ==========================================================

SECRET_KEY = os.getenv(
    "SESSION_SECRET_KEY",
    "madecentro_super_secret_2026"  # ⚠️ Cambiar en producción
)

app.add_middleware(
    SessionMiddleware,
    secret_key=SECRET_KEY,
    same_site="lax",
    https_only=False,  # 🔥 Cambiar a True cuando uses HTTPS en producción
    max_age=60 * 60 * 8  # 8 horas
)


# ==========================================================
# TEMPLATES Y STATIC
# ==========================================================

templates = Jinja2Templates(directory="templates")
app.mount("/static", StaticFiles(directory="static"), name="static")


# ==========================================================
# LOGIN MULTI PLANTA
# ==========================================================

@app.get("/login", response_class=HTMLResponse)
def login_form(request: Request):
    return templates.TemplateResponse(
        "login.html",
        {"request": request}
    )


@app.post("/login")
def login_post(request: Request, planta: str = Form(...)):

    if planta not in DATABASES:
        raise HTTPException(status_code=400, detail="Planta inválida")

    # Guardar planta en sesión
    request.session["planta_codigo"] = planta

    # Redirección segura
    response = RedirectResponse(url="/", status_code=302)
    return response


# ==========================================================
# HOME / REDIRECCIÓN SEGÚN SESIÓN
# ==========================================================

@app.get("/")
def home(request: Request):

    planta = request.session.get("planta_codigo")

    if not planta:
        return RedirectResponse(url="/login", status_code=302)

    return RedirectResponse(url="/pedidos", status_code=302)


# ==========================================================
# LOGOUT
# ==========================================================

@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/login", status_code=302)


# ==========================================================
# FUNCIÓN PARA PROTEGER RUTAS
# ==========================================================

def require_login(request: Request):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="No autenticado")

    return planta

# ==========================================================
# CONFIGURACIÓN ADMIN
# ==========================================================

PIN_ADMIN = "0813"

# ==========================================================
# CREAR CARPETAS NECESARIAS
# ==========================================================

CARPETAS_SGI = ["reportes", "remisiones", "etiquetas_pdf"]

for carpeta in CARPETAS_SGI:
    os.makedirs(carpeta, exist_ok=True)
# ==========================================================
# MODELOS PRINCIPALES SGI
# ==========================================================

# ==========================================================
# PEDIDOS
# ==========================================================

class Pedido(Base):
    __tablename__ = "pedidos"

    id = Column(Integer, primary_key=True, index=True)
    numero_pedido = Column(String(50), unique=True, index=True, nullable=False)
    cliente = Column(String(150), nullable=False)
    fecha = Column(DateTime, default=datetime.utcnow, nullable=False)

    piezas = relationship(
        "Pieza",
        back_populates="pedido",
        cascade="all, delete-orphan",
        passive_deletes=True
    )

    sesiones = relationship(
        "Sesion",
        back_populates="pedido",
        cascade="all, delete-orphan",
        passive_deletes=True
    )

    entregas_cedi = relationship(
        "EntregaCEDI",
        back_populates="pedido",
        cascade="all, delete-orphan",
        passive_deletes=True
    )


# ==========================================================
# ADMIN - ELIMINAR PEDIDO
# ==========================================================

@app.post("/admin/eliminar_pedido")
def admin_eliminar_pedido(
    request: Request,
    numero_pedido: str = Form(...),
    pin: str = Form(...),
    observacion: str = Form(...)
):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    if pin.strip() != PIN_ADMIN:
        raise HTTPException(status_code=403, detail="PIN incorrecto")

    numero_pedido = numero_pedido.strip()
    observacion = observacion.strip()

    if not numero_pedido:
        raise HTTPException(status_code=400, detail="Número de pedido vacío")

    if not observacion:
        raise HTTPException(status_code=400, detail="Observación obligatoria")

    db = get_db(planta)

    try:
        pedido = db.query(Pedido).filter(
            Pedido.numero_pedido.ilike(numero_pedido)
        ).first()

        if not pedido:
            raise HTTPException(
                status_code=404,
                detail=f"Pedido no encontrado en planta {planta}"
            )

        # 🔥 ELIMINAR DEPENDENCIAS PRIMERO (ORDEN SEGURO)
        db.query(Pieza).filter(
            Pieza.pedido_id == pedido.id
        ).delete(synchronize_session=False)

        db.query(Sesion).filter(
            Sesion.pedido_id == pedido.id
        ).delete(synchronize_session=False)

        db.query(EntregaCEDI).filter(
            EntregaCEDI.pedido_id == pedido.id
        ).delete(synchronize_session=False)

        # 🔥 ELIMINAR PEDIDO
        db.delete(pedido)

        db.commit()

        return {
            "mensaje": f"Pedido {numero_pedido} eliminado correctamente",
            "planta": planta
        }

    finally:
        db.close()
# ==========================================================
# PRODUCCIÓN - PIEZAS
# ==========================================================

class Pieza(Base):
    __tablename__ = "piezas"

    id = Column(Integer, primary_key=True, index=True)

    pedido_id = Column(
        Integer,
        ForeignKey("pedidos.id", ondelete="CASCADE"),
        nullable=False,
        index=True
    )

    # Código único visible (ej: OP-0001)
    codigo_unico = Column(
        String(100),
        unique=True,
        index=True,
        nullable=False
    )

    # 🔥 Código base normalizado a 12 dígitos para escaneo EAN
    # Este es el que debe usarse en el endpoint /escanear
    codigo_base_12 = Column(
        String(12),
        index=True,
        nullable=True
    )

    base = Column(String(50))
    altura = Column(String(50))
    canto = Column(String(50))
    servicios = Column(String(200))

    paquete = Column(String(50), index=True)

    escaneada = Column(Boolean, default=False, nullable=False)
    fecha_escaneo = Column(DateTime, nullable=True)

    pedido = relationship(
        "Pedido",
        back_populates="piezas"
    )

# ==========================================================
# SESIONES DE PRODUCCIÓN
# ==========================================================

class Sesion(Base):
    __tablename__ = "sesiones"

    id = Column(Integer, primary_key=True, index=True)
    pedido_id = Column(
        Integer,
        ForeignKey("pedidos.id", ondelete="CASCADE"),
        nullable=False,
        index=True
    )

    cedula = Column(String(20), nullable=False)
    nombre = Column(String(150), nullable=False)
    zunchadora = Column(String(100), nullable=False)

    fecha_inicio = Column(DateTime, default=datetime.utcnow, nullable=False)
    fecha_fin = Column(DateTime, nullable=True)

    pedido = relationship("Pedido", back_populates="sesiones")


# ==========================================================
# AUDITORÍA ADMINISTRATIVA
# ==========================================================

class AuditoriaAdmin(Base):
    __tablename__ = "auditoria_admin"

    id = Column(Integer, primary_key=True, index=True)
    fecha = Column(DateTime, default=datetime.utcnow, nullable=False)

    accion = Column(String(100), nullable=False)
    pedido_numero = Column(String(50), nullable=False, index=True)
    detalle = Column(Text, nullable=False)

# ==========================================================
# ENTREGA CEDI - MODELO
# ==========================================================

class EntregaCEDI(Base):
    __tablename__ = "entrega_cedi"

    id = Column(Integer, primary_key=True, index=True)

    pedido_id = Column(
        Integer,
        ForeignKey("pedidos.id", ondelete="CASCADE"),
        nullable=False,
        index=True
    )

    estado = Column(
        String(20),
        default="PENDIENTE",
        nullable=False,
        index=True
    )
    # PENDIENTE / EN_CURSO / COMPLETADO

    cedula_responsable = Column(String(20), nullable=True)
    nombre_responsable = Column(String(150), nullable=True)

    fecha_inicio = Column(DateTime, nullable=True)
    fecha_fin = Column(DateTime, nullable=True)

    paquetes_confirmados = Column(Integer, default=0, nullable=False)

    foto_remision = Column(String(300), nullable=True)
    correo_destino = Column(String(500), nullable=True)
    correo_enviado = Column(Boolean, default=False, nullable=False)

    token_validacion = Column(String(64), nullable=True, index=True)

    pedido = relationship(
        "Pedido",
        back_populates="entregas_cedi"
    )


# ==========================================================
# INICIAR ENTREGA (EMPIEZA RELOJ)
# ==========================================================

@app.post("/api/entrega_cedi/iniciar/{pedido_id}")
def iniciar_entrega(request: Request, pedido_id: int):

    planta = request.session.get("planta_codigo")
    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:
        entrega = (
            db.query(EntregaCEDI)
            .filter(EntregaCEDI.pedido_id == pedido_id)
            .order_by(EntregaCEDI.id.desc())
            .first()
        )

        if not entrega:
            raise HTTPException(status_code=404, detail="Entrega no encontrada")

        if entrega.estado != "PENDIENTE":
            raise HTTPException(status_code=400, detail="La entrega ya fue iniciada")

        entrega.estado = "EN_CURSO"
        entrega.fecha_inicio = datetime.now()
        entrega.fecha_fin = None
        entrega.paquetes_confirmados = 0

        db.commit()

        return {"mensaje": "Entrega iniciada"}

    finally:
        db.close()


# ==========================================================
# FINALIZAR CONTEO (DETENER RELOJ)
# ==========================================================

@app.post("/api/entrega_cedi/finalizar/{pedido_id}")
def finalizar_conteo(
    request: Request,
    pedido_id: int,
    paquetes_confirmados: int = Form(...)
):

    planta = request.session.get("planta_codigo")
    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:
        entrega = (
            db.query(EntregaCEDI)
            .filter(EntregaCEDI.pedido_id == pedido_id)
            .order_by(EntregaCEDI.id.desc())
            .first()
        )

        if not entrega:
            raise HTTPException(status_code=404, detail="Entrega no encontrada")

        if entrega.estado != "EN_CURSO":
            raise HTTPException(status_code=400, detail="La entrega no está en curso")

        if paquetes_confirmados <= 0:
            raise HTTPException(status_code=400, detail="Cantidad inválida")

        # 🔥 Detiene reloj
        entrega.paquetes_confirmados = paquetes_confirmados
        entrega.fecha_fin = datetime.now()

        db.commit()

        return {
            "mensaje": "Conteo confirmado correctamente",
            "fecha_fin": entrega.fecha_fin
        }

    finally:
        db.close()


# ==========================================================
# CONFIRMAR PAQUETES (MISMA LÓGICA QUE FINALIZAR)
# ==========================================================

@app.post("/api/entrega_cedi/confirmar_paquetes/{pedido_id}")
def confirmar_paquetes(
    request: Request,
    pedido_id: int,
    paquetes: int = Form(...)
):

    planta = request.session.get("planta_codigo")
    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:
        entrega = (
            db.query(EntregaCEDI)
            .filter(EntregaCEDI.pedido_id == pedido_id)
            .order_by(EntregaCEDI.id.desc())
            .first()
        )

        if not entrega:
            raise HTTPException(status_code=404, detail="Entrega no encontrada")

        if entrega.estado != "EN_CURSO":
            raise HTTPException(status_code=400, detail="La entrega no está en curso")

        if paquetes <= 0:
            raise HTTPException(status_code=400, detail="Cantidad inválida")

        # 🔥 MISMO COMPORTAMIENTO
        entrega.paquetes_confirmados = paquetes
        entrega.fecha_fin = datetime.now()

        db.commit()

        return {
            "mensaje": "Conteo confirmado",
            "fecha_fin": entrega.fecha_fin
        }

    finally:
        db.close()

# ==========================================================
# ENVIAR CORREO Y FINALIZAR (MULTIPLANTA REAL)
# ==========================================================

@app.post("/api/entrega_cedi/enviar_correo/{pedido_id}")
def enviar_correo(request: Request, pedido_id: int):

    planta = request.session.get("planta_codigo")
    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:
        entrega = (
            db.query(EntregaCEDI)
            .filter(EntregaCEDI.pedido_id == pedido_id)
            .order_by(EntregaCEDI.id.desc())
            .first()
        )

        if not entrega:
            raise HTTPException(status_code=404, detail="Entrega no encontrada")

        if entrega.estado != "EN_CURSO":
            raise HTTPException(
                status_code=400,
                detail="La entrega no está en curso"
            )

        if not entrega.paquetes_confirmados or entrega.paquetes_confirmados <= 0:
            raise HTTPException(
                status_code=400,
                detail="Debe confirmar el conteo antes de finalizar"
            )

        pedido = db.query(Pedido).filter(Pedido.id == pedido_id).first()

        if not pedido:
            raise HTTPException(status_code=404, detail="Pedido no encontrado")

        # ======================================================
        # FINALIZAR ENTREGA
        # ======================================================

        entrega.estado = "COMPLETADO"
        entrega.correo_enviado = True
        entrega.fecha_fin = datetime.now()

        # ======================================================
        # ACTUALIZAR DESPACHOS
        # ======================================================

        numero_pedido = pedido.numero_pedido.strip()

        print("OP QUE SE INTENTA MARCAR:", numero_pedido)

        op = db.query(OP).filter(OP.numero_op == numero_pedido).first()

        if not op:
            print("NO SE ENCONTRO OP EXACTA. BUSCANDO NORMALIZADA...")
            op = db.query(OP).filter(
                func.replace(OP.numero_op, " ", "") ==
                numero_pedido.replace(" ", "")
            ).first()

        if op:
            print("OP ENCONTRADA:", op.numero_op)

            op.completada = True
            op.fecha_entrega = datetime.now()

            ov = op.ov

            if ov:
                total_ops = len(ov.ops)
                completadas = sum(1 for o in ov.ops if o.completada)

                print("TOTAL OPS:", total_ops)
                print("COMPLETADAS:", completadas)

                if total_ops > 0 and completadas == total_ops:
                    ov.estado = "LISTA_PARA_DESPACHO"
                    if not ov.fecha_lista_despacho:
                        ov.fecha_lista_despacho = datetime.now()
                else:
                    ov.estado = "EN_PROCESO"

        else:
            print("NO SE ENCONTRO NINGUNA OP QUE COINCIDA")

        # ======================================================
        # 🔥 ENVÍO DE CORREO 
        # ======================================================

        try:
            numero = pedido.numero_pedido.upper()

            if numero.startswith("OV"):
                tipo_orden = "Orden de Venta (OV)"
            elif numero.startswith("OP"):
                tipo_orden = "Orden de Producción (OP)"
            else:
                tipo_orden = "Orden"

            destinatarios_to = ["steve.nova@madecentro.co"]
            destinatarios_cc = []

            if CORREOS_FIJOS:
                destinatarios_cc.extend(
                    [c.strip() for c in CORREOS_FIJOS if c.strip()]
                )

            msg = EmailMessage()
            msg["From"] = EMAIL_REMITENTE
            msg["To"] = ", ".join(destinatarios_to)

            if destinatarios_cc:
                msg["Cc"] = ", ".join(destinatarios_cc)

            msg["Subject"] = f"Entrega Oficial al CEDI - Pedido {pedido.numero_pedido}"

            msg.set_content(f"""
ENTREGA OFICIAL AL CEDI

Se informa de manera oficial que el pedido {pedido.numero_pedido},
correspondiente al cliente {pedido.cliente},
ha sido entregado al CEDI.

Tipo de orden: {tipo_orden}

Responsable de entrega:
{entrega.nombre_responsable} - {entrega.cedula_responsable}

Paquetes confirmados: {entrega.paquetes_confirmados}

Este correo es de carácter informativo y tiene como finalidad
mantener la trazabilidad de cada orden de producción
y su respectiva entrega al CEDI.

Una vez la OV se encuentre completamente entregada,
favor comunicarse con el encargado de logística
para gestionar el despacho correspondiente.

Agradecemos el compromiso y la gestión realizada.

SGI Madecentro
Sistema de Gestión Integral
""")

            # Adjuntar PDF pistoleo si existe
            if hasattr(entrega, "ruta_pdf_pistoleo") and entrega.ruta_pdf_pistoleo:
                if os.path.exists(entrega.ruta_pdf_pistoleo):
                    with open(entrega.ruta_pdf_pistoleo, "rb") as f:
                        msg.add_attachment(
                            f.read(),
                            maintype="application",
                            subtype="pdf",
                            filename=os.path.basename(entrega.ruta_pdf_pistoleo)
                        )

            # Adjuntar remisión si existe
            if hasattr(entrega, "ruta_remision") and entrega.ruta_remision:
                if os.path.exists(entrega.ruta_remision):
                    with open(entrega.ruta_remision, "rb") as f:
                        msg.add_attachment(
                            f.read(),
                            maintype="application",
                            subtype="pdf",
                            filename=os.path.basename(entrega.ruta_remision)
                        )

            with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
                server.starttls()
                server.login(EMAIL_REMITENTE, EMAIL_PASSWORD)
                server.send_message(msg)

        except Exception as correo_error:
            print("ERROR EN ENVÍO DE CORREO:", str(correo_error))

        # ======================================================
        # COMMIT FINAL 
        # ======================================================

        db.commit()

        return {
            "mensaje": "Entrega finalizada y despachos actualizado correctamente"
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        db.close()
# ==========================================================
# FUNCIONES AUXILIARES
# ==========================================================

def limpiar_paquete(paquete: str) -> str:
    if paquete is None:
        return ""
    paquete = str(paquete).strip()
    paquete = re.sub(r"\s+", " ", paquete)
    return paquete


def normalizar_paquete(paquete: str) -> str:
    """
    Normaliza el paquete para agrupar.

    Ejemplos:
    "1 | 6"  -> "1-6"
    "7 L 8"  -> "7-8"
    "9/10"   -> "9-10"
    "12"     -> "12"
    """

    if not paquete:
        return ""

    paquete = str(paquete).strip()

    # Unificar separadores como rango "-"
    paquete = paquete.replace(" | ", "-")
    paquete = paquete.replace("|", "-")
    paquete = paquete.replace("/", "-")
    paquete = paquete.replace(" L ", "-")
    paquete = paquete.replace("L", "-")
    paquete = paquete.replace(",", "-")

    # Quitar espacios
    paquete = re.sub(r"\s+", "", paquete)

    return paquete


def key_paquete_num(paquete):
    """
    Ordena por el primer número del rango.
    Ej: "7-8" -> 7
    """

    try:
        p = str(paquete).strip()

        if "-" in p:
            p = p.split("-")[0]

        return int(p)

    except Exception:
        return 999999  # envía al final valores no numéricos


def ordenar_paquetes(lista_paquetes):
    return sorted(lista_paquetes, key=key_paquete_num)


def safe_text(txt, max_len=40):
    if txt is None:
        return ""

    txt = str(txt).strip()

    if len(txt) > max_len:
        return txt[:max_len]

    return txt
# --------------------------------------------------
# NORMALIZAR CÓDIGO A 12 DÍGITOS (ESCANEO)
# --------------------------------------------------

def normalizar_codigo_12(codigo: str) -> str:
    """
    Normaliza cualquier código escaneado para comparación interna:
    - Quita caracteres no numéricos
    - Si es EAN13 (13 dígitos), elimina el dígito verificador
    - Si el lector agregó un cero al inicio y quedó >12, toma los últimos 12
    - Devuelve siempre 12 dígitos
    """
    if not codigo:
        return ""

    codigo_numerico = re.sub(r"\D", "", str(codigo))

    # Si es EAN13 (13 dígitos), quitar dígito verificador
    if len(codigo_numerico) == 13:
        codigo_numerico = codigo_numerico[:-1]

    # Si quedó mayor a 12 (por cero agregado adelante), tomar los últimos 12
    if len(codigo_numerico) > 12:
        codigo_numerico = codigo_numerico[-12:]

    # Asegurar 12 dígitos
    return codigo_numerico.zfill(12)


# ==========================================================
# HOME / PEDIDOS
# ==========================================================

@app.get("/", response_class=HTMLResponse)
def home(request: Request):

    if not request.session.get("planta_codigo"):
        return RedirectResponse(url="/login", status_code=302)

    return RedirectResponse(url="/pedidos", status_code=302)


@app.get("/pedidos", response_class=HTMLResponse)
def ver_pedidos(request: Request):

    if not request.session.get("planta_codigo"):
        return RedirectResponse(url="/login", status_code=302)

    return templates.TemplateResponse(
        "pedidos.html",
        {"request": request}
    )

# ==========================================================
# API INFO PEDIDO (PARA ENTREGA CEDI)
# ==========================================================

@app.get("/api/pedido_info/{pedido_id}")
def api_pedido_info(request: Request, pedido_id: int):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:
        pedido = db.query(Pedido).filter(
            Pedido.id == pedido_id
        ).first()

        if not pedido:
            raise HTTPException(status_code=404, detail="Pedido no encontrado")

        total = db.query(Pieza).filter(
            Pieza.pedido_id == pedido_id
        ).count()

        escaneadas = db.query(Pieza).filter(
            Pieza.pedido_id == pedido_id,
            Pieza.escaneada.is_(True)  # 🔥 mejor práctica SQLAlchemy
        ).count()

        pendientes = total - escaneadas
        porcentaje = int((escaneadas / total) * 100) if total > 0 else 0

        return {
            "id": pedido.id,
            "numero_pedido": pedido.numero_pedido,
            "cliente": pedido.cliente,
            "total": total,
            "escaneadas": escaneadas,
            "pendientes": pendientes,
            "porcentaje": porcentaje
        }

    finally:
        db.close()


# ==========================================================
# VISTA ENTREGA CEDI
# ==========================================================

@app.get("/entrega_cedi/{pedido_id}", response_class=HTMLResponse)
def entrega_cedi(request: Request, pedido_id: int):

    if not request.session.get("planta_codigo"):
        return RedirectResponse(url="/login", status_code=302)

    return templates.TemplateResponse(
        "entrega_cedi.html",
        {
            "request": request,
            "pedido_id": pedido_id
        }
    )
# ==========================================================
# API: TOTAL PAQUETES DESDE PISTOLEO
# ==========================================================

@app.get("/api/pedido_paquetes/{pedido_id}")
def api_pedido_paquetes(request: Request, pedido_id: int):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:
        # 🔎 Validar que el pedido exista
        pedido = db.query(Pedido).filter(
            Pedido.id == pedido_id
        ).first()

        if not pedido:
            raise HTTPException(status_code=404, detail="Pedido no encontrado")

        # Traer solo el campo necesario (mejor performance)
        piezas = db.query(Pieza.paquete).filter(
            Pieza.pedido_id == pedido_id
        ).all()

        paquetes = {
            normalizar_paquete(p.paquete)
            for p in piezas
            if p.paquete
        }

        return {"total_paquetes": len(paquetes)}

    finally:
        db.close()


# ==========================================================
# API: ESTADO ENTREGA
# ==========================================================

@app.get("/api/entrega_cedi/{pedido_id}")
def api_entrega_cedi(request: Request, pedido_id: int):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:
        entrega = (
            db.query(EntregaCEDI)
            .filter(EntregaCEDI.pedido_id == pedido_id)
            .order_by(EntregaCEDI.id.desc())
            .first()
        )

        # ======================================================
        # TIEMPO ACTUAL (PARA CONTROL DE DÍAS)
        # ======================================================

        ahora = datetime.utcnow()

        if not entrega:
            return {
                "existe": False,
                "estado": "PENDIENTE",
                "cedula_responsable": "",
                "nombre_responsable": "",
                "fecha_inicio": None,
                "fecha_fin": None,
                "paquetes_confirmados": 0,
                "correo_destino": None,
                "foto_remision": None,
                "correo_enviado": False,

                # ==================================================
                # CAMPOS NUEVOS (NO AFECTAN LÓGICA EXISTENTE)
                # ==================================================
                "fecha_entrada_cedi": None,
                "dias_disponible": 0,
                "semaforo": "VERDE"
            }

        # ======================================================
        # CALCULAR SEMÁFORO DE DISPONIBILIDAD
        # ======================================================

        dias = 0
        semaforo = "VERDE"

        if entrega.fecha_inicio:

            dias = (ahora - entrega.fecha_inicio).days

            if dias <= 1:
                semaforo = "VERDE"
            elif dias == 2:
                semaforo = "NARANJA"
            else:
                semaforo = "ROJO"

        return {
            "existe": True,
            "estado": entrega.estado,
            "cedula_responsable": entrega.cedula_responsable or "",
            "nombre_responsable": entrega.nombre_responsable or "",

            "fecha_inicio": entrega.fecha_inicio.strftime("%Y-%m-%d %H:%M:%S")
                if entrega.fecha_inicio else None,

            "fecha_fin": entrega.fecha_fin.strftime("%Y-%m-%d %H:%M:%S")
                if entrega.fecha_fin else None,

            "paquetes_confirmados": entrega.paquetes_confirmados or 0,
            "correo_destino": entrega.correo_destino,
            "foto_remision": entrega.foto_remision,
            "correo_enviado": bool(entrega.correo_enviado),

            # ==================================================
            # CAMPOS NUEVOS VISUALES (NO ROMPEN FRONTEND)
            # ==================================================

            "fecha_entrada_cedi": (
                (entrega.fecha_inicio - timedelta(hours=5)).strftime("%Y-%m-%d %H:%M")
                if entrega.fecha_inicio else None
            ),

            "dias_disponible": dias,
            "semaforo": semaforo
        }

    finally:
        db.close()
# ==========================================================
# INICIAR ENTREGA CEDI
# ==========================================================

@app.post("/api/entrega_cedi/iniciar/{pedido_id}")
def iniciar_entrega_cedi(
    request: Request,
    pedido_id: int,
    cedula: str = Form(...),
    nombre: str = Form(...)
):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:
        # ======================================================
        # VALIDACIONES BÁSICAS
        # ======================================================
        cedula = cedula.strip()
        nombre = nombre.strip()

        if not cedula or not nombre:
            raise HTTPException(
                status_code=400,
                detail="Debe ingresar cédula y nombre"
            )

        pedido = db.query(Pedido).filter(
            Pedido.id == pedido_id
        ).first()

        if not pedido:
            raise HTTPException(
                status_code=404,
                detail="Pedido no encontrado"
            )

        # ======================================================
        # TRAER ÚLTIMA ENTREGA
        # ======================================================
        entrega = (
            db.query(EntregaCEDI)
            .filter(EntregaCEDI.pedido_id == pedido_id)
            .order_by(EntregaCEDI.id.desc())
            .first()
        )

        # ======================================================
        # CASO 1: NO EXISTE ENTREGA → CREAR
        # ======================================================
        if not entrega:

            entrega = EntregaCEDI(
                pedido_id=pedido_id,
                estado="EN_CURSO",
                cedula_responsable=cedula,
                nombre_responsable=nombre,
                fecha_inicio=datetime.now(),
                fecha_fin=None,
                paquetes_confirmados=0,
                correo_enviado=False,
                token_validacion=None
            )

            db.add(entrega)

        # ======================================================
        # CASO 2: EXISTE → REINICIO COMPLETO CONTROLADO
        # ======================================================
        else:

            entrega.estado = "EN_CURSO"
            entrega.cedula_responsable = cedula
            entrega.nombre_responsable = nombre
            entrega.fecha_inicio = datetime.now()
            entrega.fecha_fin = None

            # 🔥 Reinicio total de variables críticas
            entrega.paquetes_confirmados = 0
            entrega.correo_enviado = False
            entrega.correo_destino = None
            entrega.token_validacion = None

        db.commit()

        return {"mensaje": "Entrega iniciada correctamente"}

    finally:
        db.close()
# ==========================================================
# SUBIR REMISIÓN
# ==========================================================

@app.post("/api/entrega_cedi/subir_remision/{pedido_id}")
async def subir_remision(
    request: Request,
    pedido_id: int,
    archivo: UploadFile = File(...)
):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:
        # 🔎 Validar que el pedido exista
        pedido = db.query(Pedido).filter(
            Pedido.id == pedido_id
        ).first()

        if not pedido:
            raise HTTPException(status_code=404, detail="Pedido no encontrado")

        # 🔎 Traer última entrega
        entrega = (
            db.query(EntregaCEDI)
            .filter(EntregaCEDI.pedido_id == pedido_id)
            .order_by(EntregaCEDI.id.desc())
            .first()
        )

        if not entrega or entrega.estado != "EN_CURSO":
            raise HTTPException(
                status_code=400,
                detail="La entrega no está en curso"
            )

        # 🔎 Validar archivo
        if not archivo.filename:
            raise HTTPException(status_code=400, detail="Archivo inválido")

        ext = os.path.splitext(archivo.filename)[1].lower()

        if ext not in [".jpg", ".jpeg", ".png", ".webp"]:
            raise HTTPException(status_code=400, detail="Formato inválido")

        # ======================================================
        # GUARDAR ARCHIVO
        # ======================================================

        os.makedirs("remisiones", exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        nombre = f"REMISION_{pedido_id}_{timestamp}{ext}"
        ruta = os.path.join("remisiones", nombre)

        with open(ruta, "wb") as buffer:
            shutil.copyfileobj(archivo.file, buffer)

        entrega.foto_remision = ruta
        db.commit()

        return {"mensaje": "Remisión cargada correctamente"}

    finally:
        db.close()
# ==========================================================
# ENVIAR CORREO Y CERRAR ENTREGA
# ==========================================================

@app.post("/api/entrega_cedi/enviar_correo/{pedido_id}")
def enviar_correo_entrega(
    request: Request,
    pedido_id: int,
    correo_destino: str = Form("")
):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:
        # ======================================================
        # VALIDACIONES BASE
        # ======================================================

        entrega = (
            db.query(EntregaCEDI)
            .filter(EntregaCEDI.pedido_id == pedido_id)
            .order_by(EntregaCEDI.id.desc())
            .first()
        )

        if not entrega:
            raise HTTPException(status_code=404, detail="Entrega no encontrada")

        if entrega.estado == "COMPLETADO":
            raise HTTPException(
                status_code=400,
                detail="La entrega ya fue finalizada"
            )

        if not entrega.paquetes_confirmados:
            raise HTTPException(status_code=400, detail="Debe confirmar el conteo antes")

        if not entrega.foto_remision or not os.path.exists(entrega.foto_remision):
            raise HTTPException(status_code=400, detail="Debe subir una remisión válida antes")

        pedido = db.query(Pedido).filter(
            Pedido.id == pedido_id
        ).first()

        if not pedido:
            raise HTTPException(status_code=404, detail="Pedido no encontrado")

        # ======================================================
        # BUSCAR PDF GENERADO
        # ======================================================

        carpeta = "reportes"

        if not os.path.exists(carpeta):
            raise HTTPException(status_code=400, detail="No existe carpeta de reportes")

        archivos = [
            f for f in os.listdir(carpeta)
            if f.startswith(f"REPORTE_COMPLETO_{pedido.numero_pedido}_")
            and f.endswith(".pdf")
        ]

        if not archivos:
            raise HTTPException(status_code=400, detail="Debe generar el PDF antes de enviar el correo")

        archivos.sort(reverse=True)
        pdf_path = os.path.join(carpeta, archivos[0])

        if not os.path.exists(pdf_path):
            raise HTTPException(status_code=400, detail="El PDF generado no existe")

        # ======================================================
        # TIPO DE ORDEN
        # ======================================================

        numero = pedido.numero_pedido.upper()

        if numero.startswith("OV"):
            tipo_orden = "Orden de Venta (OV)"
        elif numero.startswith("OP"):
            tipo_orden = "Orden de Producción (OP)"
        else:
            tipo_orden = "Orden"

        # ======================================================
        # DESTINATARIOS
        # ======================================================

        destinatarios_to = {"steve.nova@madecentro.co"}
        destinatarios_cc = set()

        if correo_destino.strip():
            manuales = {
                c.strip()
                for c in correo_destino.split(",")
                if c.strip()
            }
            destinatarios_to.update(manuales)

        if CORREOS_FIJOS:
            destinatarios_cc.update(
                {c.strip() for c in CORREOS_FIJOS if c.strip()}
            )

        if not destinatarios_to and not destinatarios_cc:
            raise HTTPException(status_code=400, detail="No hay destinatarios configurados")

        # ======================================================
        # CREAR MENSAJE
        # ======================================================

        msg = EmailMessage()
        msg["Subject"] = f"Entrega Oficial al CEDI - Pedido {pedido.numero_pedido}"
        msg["From"] = EMAIL_REMITENTE
        msg["To"] = ", ".join(destinatarios_to)

        if destinatarios_cc:
            msg["Cc"] = ", ".join(destinatarios_cc)

        msg.set_content(f"""
ENTREGA OFICIAL AL CEDI

Se informa que el pedido {pedido.numero_pedido},
cliente {pedido.cliente},
ha sido entregado al CEDI.

Tipo de orden: {tipo_orden}

Responsable:
{entrega.nombre_responsable} - {entrega.cedula_responsable}

Paquetes confirmados: {entrega.paquetes_confirmados}

SGI Madecentro
""")

        # ======================================================
        # ADJUNTAR PDF
        # ======================================================

        with open(pdf_path, "rb") as f:
            msg.add_attachment(
                f.read(),
                maintype="application",
                subtype="pdf",
                filename=os.path.basename(pdf_path)
            )

        # ======================================================
        # ADJUNTAR REMISIÓN
        # ======================================================

        ext = os.path.splitext(entrega.foto_remision)[1].lower()
        subtype_map = {
            ".jpg": "jpeg",
            ".jpeg": "jpeg",
            ".png": "png",
            ".webp": "webp"
        }

        subtype = subtype_map.get(ext, "jpeg")

        with open(entrega.foto_remision, "rb") as f:
            msg.add_attachment(
                f.read(),
                maintype="image",
                subtype=subtype,
                filename=os.path.basename(entrega.foto_remision)
            )

        # ======================================================
        # ENVIAR CORREO
        # ======================================================

        try:
            with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
                server.starttls()
                server.login(EMAIL_REMITENTE, EMAIL_PASSWORD)
                server.send_message(msg)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error enviando correo: {str(e)}")

        # ======================================================
        # CERRAR ENTREGA
        # ======================================================

        entrega.estado = "COMPLETADO"
        entrega.correo_enviado = True
        entrega.correo_destino = ", ".join(destinatarios_to.union(destinatarios_cc))
        entrega.fecha_fin = datetime.now()

        # ======================================================
        # ACTUALIZAR DESPACHOS (AJUSTE NECESARIO)
        # ======================================================

        try:
            numero_objetivo = pedido.numero_pedido.strip()

            ops_db = db.query(OP).all()
            op = None

            for o in ops_db:
                if o.numero_op and o.numero_op.strip() == numero_objetivo:
                    op = o
                    break

            if op:

                op.completada = True
                op.fecha_entrega = datetime.now()

                ov = op.ov

                total_ops = len(ov.ops)
                completadas = sum(1 for o in ov.ops if o.completada)

                if total_ops > 0 and completadas == total_ops:
                    ov.estado = "LISTA_PARA_DESPACHO"
                    if not ov.fecha_lista_despacho:
                        ov.fecha_lista_despacho = datetime.now()
                else:
                    ov.estado = "EN_PROCESO"

        except Exception as e:
            print("Error actualizando despachos:", e)

        db.commit()

        return {
            "mensaje": "Correo enviado correctamente y entrega completada"
        }

    finally:
        db.close()
# ==========================================================
# REINICIAR ENTREGA (SUPERVISOR)
# ==========================================================

@app.post("/api/entrega_cedi/reiniciar/{pedido_id}")
def reiniciar_entrega_cedi(
    request: Request,
    pedido_id: int,
    pin: str = Form(...),
    observacion: str = Form(...)
):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:
        pin = pin.strip()
        observacion = observacion.strip()

        # ======================================================
        # VALIDACIONES
        # ======================================================

        if pin != PIN_ADMIN:
            raise HTTPException(status_code=403, detail="PIN incorrecto")

        if not observacion:
            raise HTTPException(status_code=400, detail="Observación obligatoria")

        pedido = db.query(Pedido).filter(
            Pedido.id == pedido_id
        ).first()

        if not pedido:
            raise HTTPException(status_code=404, detail="Pedido no encontrado")

        entrega = (
            db.query(EntregaCEDI)
            .filter(EntregaCEDI.pedido_id == pedido_id)
            .order_by(EntregaCEDI.id.desc())
            .first()
        )

        if not entrega:
            raise HTTPException(status_code=404, detail="No existe entrega")

        # ======================================================
        # REINICIO CONTROLADO
        # ======================================================

        entrega.paquetes_confirmados = 0
        entrega.fecha_inicio = None
        entrega.fecha_fin = None
        entrega.correo_enviado = False
        entrega.correo_destino = None
        entrega.token_validacion = None
        entrega.estado = "PENDIENTE"

        # ======================================================
        # AUDITORÍA ADMINISTRATIVA
        # ======================================================

        auditoria = AuditoriaAdmin(
            accion="REINICIO_ENTREGA_CEDI",
            pedido_numero=pedido.numero_pedido,
            detalle=f"Reinicio manual por supervisor. Observación: {observacion}"
        )

        db.add(auditoria)

        db.commit()

        return {"mensaje": "Entrega reiniciada correctamente"}

    finally:
        db.close()
# ==========================================================
# SUBIR EXCEL (CORREGIDO CON codigo_base_12)
# ==========================================================

@app.post("/subir_excel")
async def subir_excel(
    request: Request,
    file: UploadFile = File(...)
):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:
        # =================================================
        # VALIDAR ARCHIVO
        # =================================================

        if not file.filename.lower().endswith((".xlsx", ".xls")):
            raise HTTPException(status_code=400, detail="Formato de archivo no permitido")

        df = pd.read_excel(file.file)

        if df.empty:
            raise HTTPException(status_code=400, detail="El archivo Excel está vacío")

        # Normalizar nombres de columnas
        df.columns = [str(c).strip().lower() for c in df.columns]

        pedido_numero = str(df.iloc[0].get("pedido", "")).strip()
        cliente = str(df.iloc[0].get("cliente", "")).strip()

        if not pedido_numero or pedido_numero.lower() == "nan":
            raise HTTPException(
                status_code=400,
                detail="El Excel no tiene número de pedido válido"
            )

        if not cliente or cliente.lower() == "nan":
            cliente = "SIN CLIENTE"

        # =================================================
        # BUSCAR PEDIDO EXISTENTE
        # =================================================

        pedido_actual = db.query(Pedido).filter(
            Pedido.numero_pedido == pedido_numero
        ).first()

        # =================================================
        # BLOQUEAR RECARGA
        # =================================================

        if pedido_actual:

            piezas_existentes = db.query(Pieza).filter(
                Pieza.pedido_id == pedido_actual.id
            ).count()

            if piezas_existentes > 0:
                raise HTTPException(
                    status_code=400,
                    detail=f"El pedido {pedido_numero} ya fue cargado. No se permite recargar."
                )

        else:
            pedido_actual = Pedido(
                numero_pedido=pedido_numero,
                cliente=cliente
            )
            db.add(pedido_actual)
            db.commit()
            db.refresh(pedido_actual)

        # =================================================
        # PIEZAS EXISTENTES
        # =================================================

        piezas_existentes = db.query(Pieza).filter(
            Pieza.pedido_id == pedido_actual.id
        ).all()

        existentes_set = {p.codigo_unico for p in piezas_existentes}
        contador = len(piezas_existentes) + 1

        # =================================================
        # RECORRER EXCEL
        # =================================================

        for _, row in df.iterrows():

            cantidad_raw = row.get("cantidad", 0)

            if pd.isna(cantidad_raw):
                continue

            try:
                cantidad = int(cantidad_raw)
            except (ValueError, TypeError):
                continue

            if cantidad <= 0:
                continue

            paquete_raw = limpiar_paquete(row.get("paquete", ""))

            # Expandir rangos tipo 1-5
            if "-" in paquete_raw:
                try:
                    inicio, fin = paquete_raw.split("-")
                    inicio = int(inicio.strip())
                    fin = int(fin.strip())
                    paquetes_expandidos = [
                        str(i) for i in range(inicio, fin + 1)
                    ]
                except Exception:
                    paquetes_expandidos = [paquete_raw]
            else:
                paquetes_expandidos = [paquete_raw]

            largo = str(row.get("largo", "")).strip()
            ancho = str(row.get("ancho", "")).strip()
            detalle = str(row.get("detalle", "")).strip()

            for i in range(cantidad):

                paquete = paquetes_expandidos[i % len(paquetes_expandidos)]
                codigo_generado = f"{pedido_numero}-{contador:04d}"

                if codigo_generado not in existentes_set:

                    # =================================================
                    # AJUSTE: USAR CÓDIGO DEL EXCEL SI EXISTE
                    # =================================================
                    codigo_excel = str(row.get("codigo", "")).strip()

                    if codigo_excel and codigo_excel.lower() != "nan":
                        codigo_base = normalizar_codigo_12(codigo_excel)
                    else:
                        codigo_base = normalizar_codigo_12(codigo_generado)

                    nueva_pieza = Pieza(
                        pedido_id=pedido_actual.id,
                        codigo_unico=codigo_generado,
                        codigo_base_12=codigo_base,
                        base=largo,
                        altura=ancho,
                        canto="",
                        servicios=detalle,
                        paquete=paquete,
                        escaneada=False
                    )

                    db.add(nueva_pieza)
                    existentes_set.add(codigo_generado)

                contador += 1

        db.commit()

        return {"mensaje": f"Pedido {pedido_numero} cargado correctamente"}

    except HTTPException:
        raise

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando Excel: {str(e)}"
        )

    finally:
        db.close()

# ==========================================================
# API PEDIDOS - PRODUCCIÓN
# ==========================================================

@app.get("/api/pedidos_produccion")
def api_pedidos_produccion(
    request: Request,
    page: int = Query(1, ge=1),
    size: int = Query(8, ge=1, le=50)
):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:
        # ======================================================
        # TRAER PEDIDOS
        # ======================================================

        pedidos = db.query(Pedido).order_by(
            Pedido.id.desc()
        ).all()

        resultado = []

        # 🔵 TIEMPO ACTUAL EN UTC (para que coincida con lo guardado en BD)
        ahora = datetime.utcnow()

        for p in pedidos:

            # 🔥 SOLO TRAER LO NECESARIO (no objetos completos)
            piezas = db.query(
                Pieza.escaneada
            ).filter(
                Pieza.pedido_id == p.id
            ).all()

            total = len(piezas)
            escaneadas = sum(1 for x in piezas if x.escaneada)

            porcentaje = int((escaneadas / total) * 100) if total > 0 else 0

            # ======================================================
            # 🔥 CONTROL 24 HORAS DESDE CARGUE
            # ======================================================

            fecha_cargue = p.fecha

            horas_desde_cargue = 0
            alerta_24h = "OK"
            semaforo_tiempo = "VERDE"

            if fecha_cargue:

                horas_desde_cargue = (
                    ahora - fecha_cargue
                ).total_seconds() / 3600

                if horas_desde_cargue < 12:
                    semaforo_tiempo = "VERDE"
                elif horas_desde_cargue < 24:
                    semaforo_tiempo = "AMARILLO"
                else:
                    semaforo_tiempo = "ROJO"

                if horas_desde_cargue > 24 and porcentaje < 100:
                    alerta_24h = "ATRASADO"

            # SOLO PRODUCCIÓN (no cumplidos)
            if porcentaje < 100:

                resultado.append({
                    "id": p.id,
                    "numero_pedido": p.numero_pedido,
                    "cliente": p.cliente,
                    "total_piezas": total,
                    "escaneadas": escaneadas,
                    "porcentaje": porcentaje,
                    "semaforo": "EN CURSO" if porcentaje > 0 else "PENDIENTE",

                    # ==================================================
                    # 🔥 CAMPOS NUEVOS PARA CONTROL OPERATIVO
                    # ==================================================

                    # 🔧 MOSTRAR HORA COLOMBIA (-5)
                    "fecha_cargue": (
                        (fecha_cargue - timedelta(hours=5)).strftime("%Y-%m-%d %H:%M")
                        if fecha_cargue else None
                    ),

                    "horas_desde_cargue": round(horas_desde_cargue, 2),
                    "alerta_24h": alerta_24h,
                    "semaforo_tiempo": semaforo_tiempo
                })

        # ======================================================
        # PAGINACIÓN
        # ======================================================

        total_registros = len(resultado)
        total_paginas = (total_registros + size - 1) // size or 1

        inicio = (page - 1) * size
        fin = inicio + size

        data_paginada = resultado[inicio:fin]

        return {
            "data": data_paginada,
            "page": page,
            "size": size,
            "total_registros": total_registros,
            "total_paginas": total_paginas
        }

    finally:
        db.close()


# ==========================================================
# API PEDIDOS - ENTREGA CEDI
# ==========================================================

@app.get("/api/pedidos_entrega")
def api_pedidos_entrega(
    request: Request,
    page: int = Query(1, ge=1),
    size: int = Query(8, ge=1, le=50)
):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:
        # ======================================================
        # TRAER ENTREGAS ACTIVAS (NO ENVIADAS)
        # ======================================================

        entregas = (
            db.query(EntregaCEDI)
            .filter(EntregaCEDI.correo_enviado.is_(False))
            .order_by(EntregaCEDI.id.desc())
            .all()
        )

        if not entregas:
            return {
                "data": [],
                "page": page,
                "size": size,
                "total_registros": 0,
                "total_paginas": 1,

                "panel_semaforo": {
                    "verde": 0,
                    "naranja": 0,
                    "rojo": 0
                }
            }

        # ======================================================
        # OBTENER PEDIDOS ÚNICOS
        # ======================================================

        pedido_ids = list({e.pedido_id for e in entregas})

        # pedidos que ya tienen entrega iniciada
        pedidos_entrega = db.query(Pedido).filter(
            Pedido.id.in_(pedido_ids)
        ).all()

        # pedidos que ya terminaron producción
        pedidos_completados = db.query(Pedido).filter(
            Pedido.estado == "COMPLETADO"
        ).all()

        # unir ambos sin duplicar
        pedidos_dict = {p.id: p for p in pedidos_entrega}

        for p in pedidos_completados:
            if p.id not in pedidos_dict:
                pedidos_dict[p.id] = p

        ahora = datetime.now()

        contador_verde = 0
        contador_naranja = 0
        contador_rojo = 0

        pedidos_agregados = set()
        resultado = []

        for entrega in entregas:
            pedido = pedidos_dict.get(entrega.pedido_id)

            if pedido and pedido.id not in pedidos_agregados:

                # ==================================================
                # FECHA ENTRADA A CEDI
                # ==================================================

                fecha_entrada = entrega.fecha_inicio or pedido.fecha

                # 🔥 SI EL PEDIDO YA ESTÁ COMPLETADO Y NO TIENE FECHA
                # REGISTRAR AUTOMÁTICAMENTE LA FECHA DE INGRESO A CEDI
                if pedido.estado == "COMPLETADO" and not entrega.fecha_inicio:
                    entrega.fecha_inicio = datetime.utcnow()
                    db.commit()
                    fecha_entrada = entrega.fecha_inicio

                dias = 0
                semaforo = "VERDE"

                if fecha_entrada:

                    dias = (ahora - fecha_entrada).days

                    if dias <= 1:
                        semaforo = "VERDE"
                        contador_verde += 1

                    elif dias == 2:
                        semaforo = "NARANJA"
                        contador_naranja += 1

                    else:
                        semaforo = "ROJO"
                        contador_rojo += 1

                resultado.append({
                    "id": pedido.id,
                    "numero_pedido": pedido.numero_pedido,
                    "cliente": pedido.cliente,

                    "fecha_entrada_cedi": (
                        fecha_entrada.strftime("%Y-%m-%d %H:%M")
                        if fecha_entrada else None
                    ),

                    "dias_disponible": dias,
                    "semaforo": semaforo
                })

                pedidos_agregados.add(pedido.id)

        # ======================================================
        # PAGINACIÓN
        # ======================================================

        total_registros = len(resultado)
        total_paginas = (total_registros + size - 1) // size or 1

        inicio = (page - 1) * size
        fin = inicio + size

        return {
            "data": resultado[inicio:fin],
            "page": page,
            "size": size,
            "total_registros": total_registros,
            "total_paginas": total_paginas,

            "panel_semaforo": {
                "verde": contador_verde,
                "naranja": contador_naranja,
                "rojo": contador_rojo
            }
        }

    finally:
        db.close()
# ==========================================================
# ===================== DESPACHOS COMPLETO =================
# ==========================================================

# ==========================================================
# MODELOS DESPACHOS
# ==========================================================

class OV(Base):
    __tablename__ = "ovs"

    id = Column(Integer, primary_key=True, index=True)
    numero_ov = Column(String(50), unique=True, index=True, nullable=False)

    estado = Column(
        String(30),
        default="EN_PROCESO",
        nullable=False,
        index=True
    )

    activa = Column(Boolean, default=True, nullable=False)

    fecha_creacion = Column(DateTime, default=datetime.utcnow, nullable=False)
    fecha_lista_despacho = Column(DateTime, nullable=True)
    fecha_despacho = Column(DateTime, nullable=True)

    ops = relationship(
        "OP",
        back_populates="ov",
        cascade="all, delete-orphan",
        passive_deletes=True
    )


class OP(Base):
    __tablename__ = "ops"

    __table_args__ = (
        UniqueConstraint("ov_id", "numero_op", name="uq_ov_op"),
    )

    id = Column(Integer, primary_key=True, index=True)

    ov_id = Column(
        Integer,
        ForeignKey("ovs.id", ondelete="CASCADE"),
        nullable=False,
        index=True
    )

    numero_op = Column(String(50), index=True, nullable=False)

    completada = Column(Boolean, default=False, nullable=False)
    fecha_entrega = Column(DateTime, nullable=True)

    ov = relationship(
        "OV",
        back_populates="ops"
    )


# ==========================================================
# VISTA DESPACHOS
# ==========================================================

@app.get("/despachos", response_class=HTMLResponse)
def vista_despachos(request: Request):

    if not request.session.get("planta_codigo"):
        return RedirectResponse(url="/login", status_code=302)

    return templates.TemplateResponse(
        "despachos.html",
        {"request": request}
    )


# ==========================================================
# SUBIR EXCEL DESPACHOS (BASE POR PLANTA REAL)
# ==========================================================

@app.post("/api/despachos/subir_base")
async def subir_excel_despachos(
    request: Request,
    file: UploadFile = File(...)
):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    # 🔥 DESPACHOS USA BASE DE LA PLANTA LOGUEADA
    db = get_db(planta)

    try:
        df = pd.read_excel(file.file)

        if df.empty:
            raise HTTPException(400, "El Excel está vacío")

        df.columns = [str(c).strip().lower() for c in df.columns]

        if "ov" not in df.columns or "op" not in df.columns:
            raise HTTPException(400, "Debe contener columnas 'ov' y 'op'")

        nuevas_ovs = 0
        nuevas_ops = 0

        for _, row in df.iterrows():

            numero_ov = str(row["ov"]).strip()
            numero_op = str(row["op"]).strip()

            if not numero_ov or numero_ov.lower() == "nan":
                continue

            if not numero_op or numero_op.lower() == "nan":
                continue

            # ==========================
            # BUSCAR O CREAR OV
            # ==========================

            ov = db.query(OV).filter(
                func.trim(OV.numero_ov) == numero_ov
            ).first()

            if not ov:
                ov = OV(numero_ov=numero_ov)
                db.add(ov)
                db.flush()
                nuevas_ovs += 1

            # ==========================
            # BUSCAR OP EN ESA OV
            # ==========================

            op = db.query(OP).filter(
                OP.ov_id == ov.id,
                func.trim(OP.numero_op) == numero_op
            ).first()

            if not op:
                db.add(
                    OP(
                        ov_id=ov.id,
                        numero_op=numero_op
                    )
                )
                nuevas_ops += 1

        db.commit()

        return {
            "mensaje": f"Base cargada correctamente en planta {planta}",
            "ovs_creadas": nuevas_ovs,
            "ops_creadas": nuevas_ops
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        db.close()
# ==========================================================
# RESUMEN DESPACHOS (BASE POR PLANTA REAL)
# ==========================================================

@app.get("/api/despachos_resumen")
def api_despachos_resumen(request: Request):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    # 🔥 DESPACHOS USA BASE DE LA PLANTA LOGUEADA
    db = get_db(planta)

    try:
        ovs = db.query(OV).filter(
            OV.activa.is_(True)
        ).all()

        total_ovs = 0
        total_ops = 0
        ovs_listas = 0
        ops_entregadas = 0

        verde = amarillo = rojo = 0

        detalle = []
        hubo_cambios = False

        for ov in ovs:

            ops = ov.ops
            total = len(ops)
            entregadas = sum(1 for o in ops if o.completada)

            porcentaje = int((entregadas / total) * 100) if total > 0 else 0

            # ======================================================
            # ACTUALIZAR ESTADO AUTOMÁTICO
            # ======================================================

            if total > 0 and entregadas == total:

                if ov.estado != "LISTA_PARA_DESPACHO":
                    ov.estado = "LISTA_PARA_DESPACHO"
                    hubo_cambios = True

                if not ov.fecha_lista_despacho:
                    ov.fecha_lista_despacho = datetime.now()
                    hubo_cambios = True

                ovs_listas += 1

            else:
                if ov.estado != "EN_PROCESO":
                    ov.estado = "EN_PROCESO"
                    hubo_cambios = True

            # ======================================================
            # CÁLCULO DÍAS EN CEDI
            # ======================================================

            dias = 0

            if ov.estado == "LISTA_PARA_DESPACHO" and ov.fecha_lista_despacho:

                base = ov.fecha_despacho or datetime.now()
                dias = (base - ov.fecha_lista_despacho).days

                if dias <= 2:
                    verde += 1
                elif dias <= 4:
                    amarillo += 1
                else:
                    rojo += 1

            total_ovs += 1
            total_ops += total
            ops_entregadas += entregadas

            detalle.append({
                "id": ov.id,
                "ov": ov.numero_ov,
                "total_ops": total,
                "entregadas": entregadas,
                "porcentaje": porcentaje,
                "estado": ov.estado,
                "dias": dias,
                "ops": [
                    {
                        "numero": o.numero_op,
                        "estado": "ENTREGADA" if o.completada else "PENDIENTE",
                        "fecha": (
                            o.fecha_entrega.strftime("%Y-%m-%d %H:%M")
                            if o.fecha_entrega else None
                        )
                    }
                    for o in ops
                ]
            })

        if hubo_cambios:
            db.commit()

        return {
            "resumen_general": {
                "total_ovs": total_ovs,
                "total_ops": total_ops,
                "ovs_listas": ovs_listas,
                "ops_entregadas": ops_entregadas
            },
            "alertas": {
                "2_dias": verde,
                "4_dias": amarillo,
                "6_mas": rojo
            },
            "detalle": detalle
        }

    finally:
        db.close()
# ==========================================================
# ANULAR OV CON PIN SUPERVISOR
# ==========================================================

@app.post("/api/despachos/anular/{ov_id}")
def anular_ov(
    request: Request,
    ov_id: int,
    pin: str = Form(...),
    observacion: str = Form(...)
):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:
        pin = pin.strip()
        observacion = observacion.strip()

        # 🔐 Validar PIN
        if pin != PIN_ADMIN:
            raise HTTPException(status_code=403, detail="PIN supervisor incorrecto")

        if not observacion:
            raise HTTPException(
                status_code=400,
                detail="Debe ingresar observación obligatoria"
            )

        ov = db.query(OV).filter(OV.id == ov_id).first()

        if not ov:
            raise HTTPException(status_code=404, detail="OV no encontrada")

        if not ov.activa:
            raise HTTPException(status_code=400, detail="La OV ya está anulada o inactiva")

        if ov.estado == "DESPACHADA":
            raise HTTPException(
                status_code=400,
                detail="No se puede anular una OV ya despachada"
            )

        # ======================================================
        # ANULAR OV
        # ======================================================

        ov.activa = False
        ov.estado = "ANULADA"

        # Auditoría administrativa
        db.add(AuditoriaAdmin(
            accion="ANULAR_OV",
            pedido_numero=ov.numero_ov,
            detalle=f"OV anulada por supervisor. Observación: {observacion}"
        ))

        db.commit()

        return {"mensaje": "OV anulada correctamente"}

    finally:
        db.close()


# ==========================================================
# DESPACHAR OV
# ==========================================================

@app.post("/api/despachos/despachar/{ov_id}")
def despachar_ov(
    request: Request,
    ov_id: int
):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:
        ov = db.query(OV).filter(OV.id == ov_id).first()

        if not ov:
            raise HTTPException(status_code=404, detail="OV no encontrada")

        if not ov.activa:
            raise HTTPException(
                status_code=400,
                detail="No se puede despachar una OV inactiva"
            )

        if ov.estado != "LISTA_PARA_DESPACHO":
            raise HTTPException(
                status_code=400,
                detail="OV no está lista para despacho"
            )

        # ======================================================
        # MARCAR COMO DESPACHADA
        # ======================================================

        ov.estado = "DESPACHADA"
        ov.fecha_despacho = datetime.now()
        ov.activa = False

        # Auditoría
        db.add(AuditoriaAdmin(
            accion="DESPACHAR_OV",
            pedido_numero=ov.numero_ov,
            detalle="OV marcada como despachada"
        ))

        db.commit()

        return {"mensaje": "OV despachada correctamente"}

    finally:
        db.close()
# ==========================================================
# INICIAR SESIÓN
# ==========================================================

@app.get("/iniciar_sesion/{pedido_id}", response_class=HTMLResponse)
def iniciar_sesion(request: Request, pedido_id: int):

    if not request.session.get("planta_codigo"):
        return RedirectResponse(url="/login", status_code=302)

    return templates.TemplateResponse(
        "iniciar_sesion.html",
        {
            "request": request,
            "pedido_id": pedido_id
        }
    )


class DatosSesion(BaseModel):
    cedula: str
    nombre: str
    zunchadora: str


@app.post("/crear_sesion/{pedido_id}")
def crear_sesion(
    request: Request,
    pedido_id: int,
    datos: DatosSesion
):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:
        pedido = db.query(Pedido).filter(
            Pedido.id == pedido_id
        ).first()

        if not pedido:
            raise HTTPException(status_code=404, detail="Pedido no encontrado")

        cedula = str(datos.cedula).strip()
        nombre = str(datos.nombre).strip()
        zunchadora = str(datos.zunchadora).strip()

        if (
            not cedula or cedula.lower() == "nan" or
            not nombre or nombre.lower() == "nan" or
            not zunchadora or zunchadora.lower() == "nan"
        ):
            raise HTTPException(
                status_code=400,
                detail="Debe completar cédula, nombre y zunchadora"
            )

        # ======================================================
        # CREAR SESIÓN
        # ======================================================

        nueva_sesion = Sesion(
            pedido_id=pedido_id,
            cedula=cedula,
            nombre=nombre,
            zunchadora=zunchadora,
            fecha_inicio=datetime.now()
        )

        db.add(nueva_sesion)
        db.commit()
        db.refresh(nueva_sesion)

        return {"sesion_id": nueva_sesion.id}

    finally:
        db.close()
# ==========================================================
# PISTOLEO
# ==========================================================

@app.get("/pistoleo/{pedido_id}/{sesion_id}", response_class=HTMLResponse)
def pistoleo(request: Request, pedido_id: int, sesion_id: int):

    planta = request.session.get("planta_codigo")

    if not planta:
        return RedirectResponse(url="/login", status_code=302)

    db = get_db(planta)

    try:
        pedido = db.query(Pedido).filter(
            Pedido.id == pedido_id
        ).first()

        if not pedido:
            raise HTTPException(status_code=404, detail="Pedido no encontrado")

        sesion = db.query(Sesion).filter(
            Sesion.id == sesion_id,
            Sesion.pedido_id == pedido_id
        ).first()

        if not sesion:
            raise HTTPException(status_code=404, detail="Sesión no encontrada")

        # ======================================================
        # VALIDACIONES ADICIONALES DE SEGURIDAD
        # ======================================================

        if sesion.fecha_fin:
            raise HTTPException(
                status_code=400,
                detail="La sesión ya fue cerrada"
            )

        return templates.TemplateResponse(
            "pistoleo.html",
            {
                "request": request,
                "pedido_id": pedido_id,
                "sesion_id": sesion_id
            }
        )

    finally:
        db.close()
# ==========================================================
# FONDO CORPORATIVO MADECENTRO
# ==========================================================

from reportlab.lib import colors, pagesizes
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.graphics.barcode import qr
from reportlab.graphics.shapes import Drawing
from reportlab.graphics import renderPDF
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics

def fondo_corporativo(canvas, doc):
    width, height = pagesizes.letter

    # Marco naranja
    canvas.setStrokeColor(colors.HexColor("#f37021"))
    canvas.setLineWidth(3)
    canvas.rect(20, 20, width - 40, height - 40)

    # Marca de agua
    canvas.saveState()
    canvas.setFont("Helvetica-Bold", 60)
    canvas.setFillColor(colors.HexColor("#f2f2f2"))
    canvas.drawCentredString(width / 2, height / 2, "MADECENTRO")
    canvas.restoreState()


# ==========================================================
# REPORTE PDF COMPLETO ENTREGA CEDI
# ==========================================================

@app.get("/reporte_pdf/{pedido_id}")
def reporte_pdf(request: Request, pedido_id: int):

    planta = request.session.get("planta_codigo")
    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:
        pedido = db.query(Pedido).filter(Pedido.id == pedido_id).first()
        if not pedido:
            raise HTTPException(status_code=404, detail="Pedido no encontrado")

        entrega = (
            db.query(EntregaCEDI)
            .filter(EntregaCEDI.pedido_id == pedido.id)
            .order_by(EntregaCEDI.id.desc())
            .first()
        )

        if not entrega:
            raise HTTPException(status_code=404, detail="No existe registro de entrega")

        if not entrega.token_validacion:
            entrega.token_validacion = uuid.uuid4().hex
            db.commit()

        token = entrega.token_validacion

        piezas = (
            db.query(Pieza)
            .filter(Pieza.pedido_id == pedido.id)
            .order_by(Pieza.codigo_unico.asc())
            .all()
        )

        total_piezas = len(piezas)
        escaneadas = sum(1 for p in piezas if p.escaneada)
        porcentaje = round((escaneadas / total_piezas) * 100, 2) if total_piezas else 0

        estado = entrega.estado or "-"
        responsable = entrega.nombre_responsable or "-"
        cedula = entrega.cedula_responsable or "-"
        paquetes_confirmados = entrega.paquetes_confirmados or 0
        fecha_inicio = entrega.fecha_inicio.strftime("%Y-%m-%d %H:%M:%S") if entrega.fecha_inicio else "-"
        fecha_fin = entrega.fecha_fin.strftime("%Y-%m-%d %H:%M:%S") if entrega.fecha_fin else "-"
        correo_enviado = "SI" if entrega.correo_enviado else "NO"

        os.makedirs("reportes", exist_ok=True)

        numero_seguro = re.sub(r"[^A-Za-z0-9_-]", "_", pedido.numero_pedido)
        timestamp = int(datetime.now().timestamp())
        ruta_final = f"reportes/REPORTE_COMPLETO_{numero_seguro}_{timestamp}.pdf"

        doc = SimpleDocTemplate(
            ruta_final,
            pagesize=pagesizes.letter,
            rightMargin=40,
            leftMargin=40,
            topMargin=60,
            bottomMargin=60
        )

        elements = []
        styles = getSampleStyleSheet()

        # LOGO
        logo_path = "static/logo.png"
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=2.5 * inch, height=0.8 * inch)
            logo.hAlign = "CENTER"
            elements.append(logo)

        elements.append(Spacer(1, 20))

        # TITULO
        titulo_style = ParagraphStyle(
            "titulo",
            parent=styles["Heading1"],
            textColor=colors.HexColor("#f37021"),
            alignment=1
        )

        elements.append(Paragraph(
            "REPORTE OFICIAL DE CONTROL DE ESCANEO Y ENTREGA CEDI",
            titulo_style
        ))
        elements.append(Spacer(1, 25))

        # TABLA PRINCIPAL
        data_principal = [
            ["Pedido", pedido.numero_pedido],
            ["Cliente", pedido.cliente],
            ["Estado Entrega", estado],
            ["Responsable", responsable],
            ["Cédula", cedula],
        ]

        tabla_principal = Table(data_principal, colWidths=[180, 320])
        tabla_principal.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#0f1a2b")),
            ("TEXTCOLOR", (0, 0), (0, -1), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))

        elements.append(tabla_principal)
        elements.append(Spacer(1, 25))

        # RESUMEN ESCANEO
        subtitulo_style = ParagraphStyle(
            "subtitulo",
            parent=styles["Heading2"],
            textColor=colors.HexColor("#f37021")
        )

        elements.append(Paragraph("RESUMEN CONTROL DE ESCANEO", subtitulo_style))
        elements.append(Spacer(1, 12))

        data_resumen = [
            ["Total Piezas", total_piezas],
            ["Escaneadas", escaneadas],
            ["Nivel Cumplimiento", f"{porcentaje}%"],
        ]

        tabla_resumen = Table(data_resumen, colWidths=[180, 320])
        tabla_resumen.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f37021")),
            ("TEXTCOLOR", (0, 0), (0, -1), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))

        elements.append(tabla_resumen)
        elements.append(Spacer(1, 25))

        # RESUMEN ENTREGA
        elements.append(Paragraph("RESUMEN ENTREGA", styles["Heading2"]))
        elements.append(Spacer(1, 12))

        data_entrega = [
            ["Paquetes Confirmados", paquetes_confirmados],
            ["Inicio Conteo", fecha_inicio],
            ["Fin Conteo", fecha_fin],
            ["Correo Enviado", correo_enviado],
        ]

        tabla_entrega = Table(data_entrega, colWidths=[180, 320])
        tabla_entrega.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))

        elements.append(tabla_entrega)
        elements.append(Spacer(1, 30))

        # DETALLE PIEZAS
        elements.append(Paragraph("DETALLE TÉCNICO DE PIEZAS", subtitulo_style))
        elements.append(Spacer(1, 12))

        data = [["Código", "Base", "Altura", "Paquete", "Escaneada"]]

        for p in piezas:
            data.append([
                p.codigo_unico,
                getattr(p, "base", "-"),
                getattr(p, "altura", "-"),
                getattr(p, "paquete", "-"),
                "SI" if p.escaneada else "NO"
            ])

        tabla_piezas = Table(data, repeatRows=1)
        tabla_piezas.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f37021")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ]))

        elements.append(tabla_piezas)
        elements.append(Spacer(1, 30))

        # TEXTO FINAL
        elements.append(Paragraph(
            "Este documento certifica oficialmente la ejecución del proceso de escaneo, "
            "verificación operativa y entrega formal al CEDI, garantizando trazabilidad total del pedido.",
            styles["Normal"]
        ))

        elements.append(Spacer(1, 12))

        elements.append(Paragraph(
            "<b>MADECENTRO construye confianza, pieza por pieza.</b>",
            styles["Normal"]
        ))

        elements.append(Spacer(1, 8))

        elements.append(Paragraph(
            "Cada pieza validada representa compromiso, precisión y responsabilidad en cada etapa del proceso.",
            styles["Normal"]
        ))

        elements.append(Spacer(1, 18))

        elements.append(Paragraph(
            "<b>Validación Digital Oficial</b>",
            styles["Normal"]
        ))

        elements.append(Spacer(1, 12))

        qr_code = qr.QrCodeWidget(token)
        bounds = qr_code.getBounds()
        w = bounds[2] - bounds[0]
        h = bounds[3] - bounds[1]

        d = Drawing(120, 120, transform=[120.0 / w, 0, 0, 120.0 / h, 0, 0])
        d.add(qr_code)
        elements.append(d)

        doc.build(
            elements,
            onFirstPage=fondo_corporativo,
            onLaterPages=fondo_corporativo
        )

        return FileResponse(
            ruta_final,
            media_type="application/pdf",
            filename=os.path.basename(ruta_final)
        )

    finally:
        db.close()
# ==========================================================
# ESTADO PEDIDO
# ==========================================================

@app.get("/estado/{pedido_id}")
def estado(
    request: Request,
    pedido_id: int
):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:
        pedido = db.query(Pedido).filter(
            Pedido.id == pedido_id
        ).first()

        if not pedido:
            raise HTTPException(status_code=404, detail="Pedido no encontrado")

        # 🔥 Traer solo columnas necesarias
        piezas = db.query(
            Pieza.id,
            Pieza.codigo_unico,
            Pieza.base,
            Pieza.altura,
            Pieza.paquete,
            Pieza.escaneada
        ).filter(
            Pieza.pedido_id == pedido.id
        ).order_by(
            Pieza.id.asc()
        ).all()

        total = len(piezas)
        escaneadas = sum(1 for p in piezas if p.escaneada)

        paquetes_dict = {}

        for p in piezas:

            paquete_limpio = normalizar_paquete(p.paquete) or "SIN_PAQUETE"

            paquetes_dict.setdefault(paquete_limpio, []).append({
                "pieza_id": p.id,
                "codigo": p.codigo_unico,
                "largo": p.base,
                "ancho": p.altura,
                "escaneada": p.escaneada
            })

        pendientes = total - escaneadas
        porcentaje = int(
            (escaneadas / total) * 100
        ) if total > 0 else 0

        paquetes = []

        paquetes_ordenados = ordenar_paquetes(
            list(paquetes_dict.keys())
        )

        for num_paquete in paquetes_ordenados:

            lista_piezas = paquetes_dict[num_paquete]

            total_paq = len(lista_piezas)
            escaneadas_paq = sum(
                1 for x in lista_piezas if x["escaneada"]
            )

            porcentaje_paq = int(
                (escaneadas_paq / total_paq) * 100
            ) if total_paq > 0 else 0

            paquetes.append({
                "paquete": num_paquete,
                "total": total_paq,
                "escaneadas": escaneadas_paq,
                "porcentaje": porcentaje_paq,
                "piezas": lista_piezas
            })

        return {
            "pedido": pedido.numero_pedido,
            "cliente": pedido.cliente,
            "total": total,
            "escaneadas": escaneadas,
            "pendientes": pendientes,
            "porcentaje": porcentaje,
            "paquetes": paquetes
        }

    finally:
        db.close()
# ==========================================================
# ESCANEAR PIEZA (MULTIPLANTA CORREGIDO DEFINITIVO)
# ==========================================================

from sqlalchemy import func
import re

class CodigoEscaneo(BaseModel):
    codigo: str


@app.post("/escanear/{pedido_id}/{sesion_id}")
async def escanear_pieza(request: Request, pedido_id: int, sesion_id: int, datos: CodigoEscaneo):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:
        codigo_raw = str(datos.codigo).strip()

        if not codigo_raw:
            raise HTTPException(status_code=400, detail="Código vacío")

        # --------------------------------------------------
        # 1️⃣ LIMPIAR A SOLO NÚMEROS
        # --------------------------------------------------
        codigo_numerico = re.sub(r"\D", "", codigo_raw)

        # --------------------------------------------------
        # 2️⃣ SI ES EAN13 → QUITAR DÍGITO VERIFICADOR
        # --------------------------------------------------
        if len(codigo_numerico) == 13:
            codigo_numerico = codigo_numerico[:-1]

        # --------------------------------------------------
        # 3️⃣ SI ES MAYOR A 12 → TOMAR LOS ÚLTIMOS 12
        # --------------------------------------------------
        if len(codigo_numerico) > 12:
            codigo_numerico = codigo_numerico[-12:]

        # --------------------------------------------------
        # 4️⃣ NORMALIZAR A 12 DÍGITOS
        # --------------------------------------------------
        codigo_normalizado = codigo_numerico.zfill(12)

        # --------------------------------------------------
        # 5️⃣ CORRECCIÓN UPC SOLO PARA CÓDIGOS DE PLANTA
        # (los que realmente empiezan por 0: 052, 064, 050...)
        # --------------------------------------------------
        prefijos_planta_sin_cero = ("52", "64", "50")

        if codigo_numerico.startswith(prefijos_planta_sin_cero):
            if len(codigo_numerico) == 12:
                # reconstruir el cero perdido y quitar verificador falso
                codigo_normalizado = "0" + codigo_numerico[:-1]

        # --------------------------------------------------
        # VALIDAR PEDIDO
        # --------------------------------------------------
        pedido = db.query(Pedido).filter(Pedido.id == pedido_id).first()
        if not pedido:
            raise HTTPException(status_code=404, detail="Pedido no encontrado")

        # --------------------------------------------------
        # VALIDAR SESIÓN
        # --------------------------------------------------
        sesion = db.query(Sesion).filter(
            Sesion.id == sesion_id,
            Sesion.pedido_id == pedido_id
        ).first()

        if not sesion:
            raise HTTPException(status_code=404, detail="Sesión no encontrada")

        # --------------------------------------------------
        # REGLA GLOBAL (600, 700, 800, 900)
        # --------------------------------------------------
        prefijos_globales = ("600", "700", "800", "900")

        if codigo_normalizado.startswith(prefijos_globales):
            pieza = db.query(Pieza).filter(
                Pieza.codigo_base_12 == codigo_normalizado
            ).first()
        else:
            pieza = db.query(Pieza).filter(
                Pieza.pedido_id == pedido_id,
                Pieza.codigo_base_12 == codigo_normalizado
            ).first()

        # --------------------------------------------------
        # SI NO EXISTE
        # --------------------------------------------------
        if not pieza:
            raise HTTPException(
                status_code=404,
                detail=(
                    f"Pieza no encontrada.\n"
                    f"Código escaneado: {codigo_raw}\n"
                    f"Código normalizado: {codigo_normalizado}\n"
                    f"Pedido ID: {pedido_id}"
                )
            )

        if pieza.escaneada:
            raise HTTPException(status_code=400, detail="Pieza ya escaneada")

        # --------------------------------------------------
        # MARCAR COMO ESCANEADA
        # --------------------------------------------------
        pieza.escaneada = True
        pieza.fecha_escaneo = datetime.now()

        db.flush()

        # --------------------------------------------------
        # VALIDAR SI EL PEDIDO QUEDÓ COMPLETO
        # --------------------------------------------------
        piezas = db.query(Pieza).filter(
            Pieza.pedido_id == pieza.pedido_id
        ).all()

        total = len(piezas)
        escaneadas = sum(1 for p in piezas if p.escaneada)

        if total > 0 and escaneadas == total:

            entrega_existente = (
                db.query(EntregaCEDI)
                .filter(EntregaCEDI.pedido_id == pieza.pedido_id)
                .order_by(EntregaCEDI.id.desc())
                .first()
            )

            if not entrega_existente:
                nueva_entrega = EntregaCEDI(
                    pedido_id=pieza.pedido_id,
                    estado="PENDIENTE",
                    paquetes_confirmados=0,
                    fecha_inicio=None,
                    fecha_fin=None,
                    cedula_responsable=None,
                    nombre_responsable=None,
                    correo_enviado=False
                )
                db.add(nueva_entrega)

        db.commit()

        return {"mensaje": "Pieza escaneada correctamente"}

    finally:
        db.close()
# ==========================================================
# ADMIN - MODIFICAR PIEZAS / PAQUETES (PIN + OBS)
# ==========================================================

@app.post("/admin/agregar_pieza")
def admin_agregar_pieza(
    request: Request,
    pedido_id: int = Form(...),
    codigo_unico: str = Form(...),
    paquete: str = Form(""),
    largo: str = Form(""),
    ancho: str = Form(""),
    detalle: str = Form(""),
    pin: str = Form(...),
    observacion: str = Form(...)
):
    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:
        pin = pin.strip()
        observacion = observacion.strip()
        codigo_unico = codigo_unico.strip()

        if pin != PIN_ADMIN:
            raise HTTPException(status_code=403, detail="PIN incorrecto")

        if not observacion:
            raise HTTPException(status_code=400, detail="Observación obligatoria")

        if not codigo_unico:
            raise HTTPException(status_code=400, detail="Código único obligatorio")

        pedido = db.query(Pedido).filter(
            Pedido.id == pedido_id
        ).first()

        if not pedido:
            raise HTTPException(status_code=404, detail="Pedido no encontrado")

        existe = db.query(Pieza).filter(
            Pieza.codigo_unico == codigo_unico
        ).first()

        if existe:
            raise HTTPException(
                status_code=400,
                detail="Ya existe una pieza con ese código"
            )

        codigo_base = normalizar_codigo_12(codigo_unico)

        nueva = Pieza(
            pedido_id=pedido_id,
            codigo_unico=codigo_unico,
            codigo_base_12=codigo_base,
            base=largo.strip(),
            altura=ancho.strip(),
            servicios=detalle.strip(),
            paquete=limpiar_paquete(paquete),
            escaneada=False
        )

        db.add(nueva)

        db.add(AuditoriaAdmin(
            accion="AGREGAR_PIEZA",
            pedido_numero=pedido.numero_pedido,
            detalle=f"Código: {codigo_unico} | Paquete: {paquete} | Obs: {observacion}"
        ))

        db.commit()

        return {"mensaje": f"Pieza {codigo_unico} agregada correctamente"}

    finally:
        db.close()
# ==========================================================
# ADMIN - AGREGAR PAQUETE (CON PRIMERA PIEZA REAL)
# ==========================================================

@app.post("/admin/agregar_paquete")
def admin_agregar_paquete(
    request: Request,
    pedido_id: int = Form(...),
    paquete: str = Form(...),
    codigo_unico: str = Form(...),
    largo: str = Form(""),
    ancho: str = Form(""),
    detalle: str = Form(""),
    pin: str = Form(...),
    observacion: str = Form(...)
):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:
        pin = pin.strip()
        observacion = observacion.strip()
        codigo_unico = codigo_unico.strip()
        paquete = limpiar_paquete(paquete)

        if pin != PIN_ADMIN:
            raise HTTPException(status_code=403, detail="PIN incorrecto")

        if not observacion:
            raise HTTPException(status_code=400, detail="Observación obligatoria")

        if not paquete:
            raise HTTPException(status_code=400, detail="Paquete obligatorio")

        if not codigo_unico:
            raise HTTPException(status_code=400, detail="Código único obligatorio")

        pedido = db.query(Pedido).filter(
            Pedido.id == pedido_id
        ).first()

        if not pedido:
            raise HTTPException(status_code=404, detail="Pedido no encontrado")

        paquete_normalizado = normalizar_paquete(paquete)

        # Validar si ya existe paquete (sin traer todo)
        existe_paquete = db.query(Pieza.id).filter(
            Pieza.pedido_id == pedido_id
        ).filter(
            Pieza.paquete.isnot(None)
        ).all()

        for p in existe_paquete:
            pieza = db.query(Pieza.paquete).filter(Pieza.id == p.id).first()
            if pieza and normalizar_paquete(pieza.paquete) == paquete_normalizado:
                raise HTTPException(status_code=400, detail="Ese paquete ya existe")

        # Validar si ya existe el código
        existe_codigo = db.query(Pieza).filter(
            Pieza.codigo_unico == codigo_unico
        ).first()

        if existe_codigo:
            raise HTTPException(
                status_code=400,
                detail="Ya existe una pieza con ese código"
            )

        codigo_base = normalizar_codigo_12(codigo_unico)

        nueva = Pieza(
            pedido_id=pedido_id,
            codigo_unico=codigo_unico,
            codigo_base_12=codigo_base,
            base=largo.strip(),
            altura=ancho.strip(),
            canto="",
            servicios=detalle.strip(),
            paquete=paquete,
            escaneada=False
        )

        db.add(nueva)

        db.add(AuditoriaAdmin(
            accion="AGREGAR_PAQUETE",
            pedido_numero=pedido.numero_pedido,
            detalle=(
                f"Paquete agregado: {paquete} | "
                f"Primera pieza: {codigo_unico} | "
                f"Obs: {observacion}"
            )
        ))

        db.commit()

        return {
            "mensaje": (
                f"Paquete {paquete} agregado correctamente "
                f"con la pieza {codigo_unico}"
            )
        }

    finally:
        db.close()
# ==========================================================
# ADMIN - ELIMINAR PAQUETE COMPLETO
# ==========================================================

@app.post("/admin/eliminar_paquete")
def admin_eliminar_paquete(
    request: Request,
    pedido_id: int = Form(...),
    paquete: str = Form(...),
    pin: str = Form(...),
    observacion: str = Form(...)
):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:
        pin = pin.strip()
        observacion = observacion.strip()
        paquete = limpiar_paquete(paquete)

        if pin != PIN_ADMIN:
            raise HTTPException(status_code=403, detail="PIN incorrecto")

        if not observacion:
            raise HTTPException(
                status_code=400,
                detail="Observación obligatoria"
            )

        if not paquete:
            raise HTTPException(
                status_code=400,
                detail="Paquete obligatorio"
            )

        pedido = db.query(Pedido).filter(
            Pedido.id == pedido_id
        ).first()

        if not pedido:
            raise HTTPException(
                status_code=404,
                detail="Pedido no encontrado"
            )

        paquete_norm = normalizar_paquete(paquete)

        # 🔥 Traer solo piezas necesarias
        piezas = db.query(Pieza).filter(
            Pieza.pedido_id == pedido_id
        ).all()

        piezas_borrar = [
            p for p in piezas
            if normalizar_paquete(p.paquete) == paquete_norm
        ]

        if not piezas_borrar:
            raise HTTPException(
                status_code=404,
                detail="No hay piezas en ese paquete"
            )

        # 🔐 Protección: no eliminar piezas ya escaneadas
        if any(p.escaneada for p in piezas_borrar):
            raise HTTPException(
                status_code=400,
                detail="No se puede eliminar un paquete con piezas ya escaneadas"
            )

        total_eliminadas = len(piezas_borrar)

        for p in piezas_borrar:
            db.delete(p)

        db.add(AuditoriaAdmin(
            accion="ELIMINAR_PAQUETE",
            pedido_numero=pedido.numero_pedido,
            detalle=(
                f"Paquete eliminado: {paquete} | "
                f"Total piezas: {total_eliminadas} | "
                f"Obs: {observacion}"
            )
        ))

        db.commit()

        return {
            "mensaje": (
                f"Paquete {paquete} eliminado "
                f"({total_eliminadas} piezas)"
            )
        }

    finally:
        db.close() 
# ==========================================================
# DASHBOARD DIRECCIÓN SGI - VERSION EJECUTIVA COMPLETA
# ==========================================================

from collections import defaultdict
from datetime import datetime
from sqlalchemy import func


# ==========================================================
# FUNCION CENTRAL REUTILIZABLE (NO ROMPE NADA)
# ==========================================================

def obtener_metricas_dashboard(db, fecha_desde, fecha_hasta):

    # ======================================================
    # 🔵 PRODUCCIÓN
    # ======================================================
    pedidos = db.query(Pedido).all()

    pendientes = 0
    en_proceso = 0
    completadas = 0
    tendencia_produccion = defaultdict(int)
    registro_produccion = []

    for p in pedidos:

        piezas = p.piezas
        total = len(piezas)
        esc = sum(1 for x in piezas if x.escaneada)

        if total == 0:
            continue

        porc = int((esc / total) * 100)
        ultima = None  # 🔥 Reinicio seguro

        if porc == 0:
            pendientes += 1

        elif porc < 100:
            en_proceso += 1

        else:
            completadas += 1

            fechas = [x.fecha_escaneo for x in piezas if x.fecha_escaneo]

            if fechas:
                ultima = max(fechas)

                if fecha_desde <= ultima <= fecha_hasta:
                    fecha_key = ultima.strftime("%Y-%m-%d")
                    tendencia_produccion[fecha_key] += 1

        registro_produccion.append({
            "pedido": p.numero_pedido,
            "cliente": p.cliente,
            "total_piezas": total,
            "escaneadas": esc,
            "porcentaje": porc,
            "ultima_fecha": ultima.strftime("%Y-%m-%d") if ultima else None
        })

    total_ops = pendientes + en_proceso + completadas

    cumplimiento_produccion = round(
        (completadas / total_ops) * 100, 2
    ) if total_ops > 0 else 0


    # ======================================================
    # 🟢 ENTREGA CEDI
    # ======================================================
    entregas = db.query(EntregaCEDI).all()

    ent_pendientes = 0
    ent_curso = 0
    ent_completadas = 0
    tiempos = []
    tendencia_entregas = defaultdict(int)
    registro_entregas = []

    for e in entregas:

        if e.estado == "PENDIENTE":
            ent_pendientes += 1
        elif e.estado == "EN_CURSO":
            ent_curso += 1
        elif e.estado in ["COMPLETADO", "CORREO_ENVIADO"]:
            ent_completadas += 1

        minutos = None

        if e.fecha_inicio and e.fecha_fin:
            minutos = round(
                (e.fecha_fin - e.fecha_inicio).total_seconds() / 60, 2
            )
            tiempos.append(minutos)

            if fecha_desde <= e.fecha_fin <= fecha_hasta:
                fecha_key = e.fecha_fin.strftime("%Y-%m-%d")
                tendencia_entregas[fecha_key] += 1

        registro_entregas.append({
            "pedido_id": e.pedido_id,
            "responsable": e.nombre_responsable,
            "inicio": e.fecha_inicio.strftime("%Y-%m-%d %H:%M") if e.fecha_inicio else None,
            "fin": e.fecha_fin.strftime("%Y-%m-%d %H:%M") if e.fecha_fin else None,
            "minutos": minutos,
            "cumple": True if minutos is not None and minutos <= 120 else False if minutos is not None else None
        })

    total_tiempos = len(tiempos)
    cumplen_entrega = sum(1 for t in tiempos if t <= 120)

    promedio_entrega = round(sum(tiempos) / total_tiempos, 2) if total_tiempos > 0 else 0

    eficiencia_entrega = round(
        (cumplen_entrega / total_tiempos) * 100, 2
    ) if total_tiempos > 0 else 0


    # ======================================================
    # 🟠 DESPACHOS
    # ======================================================
    ovs = db.query(OV).all()

    ov_proceso = 0
    ov_lista = 0
    ov_despachada = 0
    dias_cedi = []
    tendencia_despacho = defaultdict(int)
    registro_despachos = []

    for ov in ovs:

        if ov.estado == "EN_PROCESO":
            ov_proceso += 1
        elif ov.estado == "LISTA_PARA_DESPACHO":
            ov_lista += 1
        elif ov.estado == "DESPACHADA":
            ov_despachada += 1

        dias = None

        if ov.fecha_lista_despacho and ov.fecha_despacho:
            dias = (ov.fecha_despacho - ov.fecha_lista_despacho).days
            dias_cedi.append(dias)

            if fecha_desde <= ov.fecha_despacho <= fecha_hasta:
                fecha_key = ov.fecha_despacho.strftime("%Y-%m-%d")
                tendencia_despacho[fecha_key] += 1

        registro_despachos.append({
            "ov": ov.numero_ov,
            "fecha_lista": ov.fecha_lista_despacho.strftime("%Y-%m-%d") if ov.fecha_lista_despacho else None,
            "fecha_despacho": ov.fecha_despacho.strftime("%Y-%m-%d") if ov.fecha_despacho else None,
            "dias": dias,
            "cumple": True if dias is not None and dias <= 5 else False if dias is not None else None
        })

    total_dias = len(dias_cedi)
    cumplen_despacho = sum(1 for d in dias_cedi if d <= 5)

    promedio_dias = round(sum(dias_cedi) / total_dias, 2) if total_dias > 0 else 0

    eficiencia_despacho = round(
        (cumplen_despacho / total_dias) * 100, 2
    ) if total_dias > 0 else 0


    # ======================================================
    # 📊 CONSOLIDADO DIARIO
    # ======================================================
    consolidado = defaultdict(lambda: {
        "produccion": 0,
        "entregas": 0,
        "despachos": 0
    })

    for fecha, valor in tendencia_produccion.items():
        consolidado[fecha]["produccion"] = valor

    for fecha, valor in tendencia_entregas.items():
        consolidado[fecha]["entregas"] = valor

    for fecha, valor in tendencia_despacho.items():
        consolidado[fecha]["despachos"] = valor

    consolidado_ordenado = dict(
        sorted(consolidado.items(), key=lambda x: x[0])
    )


    # ======================================================
    # 🏛 INDICE GLOBAL SGI
    # ======================================================
    indice_sgi = round(
        (cumplimiento_produccion + eficiencia_entrega + eficiencia_despacho) / 3,
        2
    )


    return {
        "indice_sgi": indice_sgi,
        "produccion": {
            "pendientes": pendientes,
            "en_proceso": en_proceso,
            "completadas": completadas,
            "cumplimiento": cumplimiento_produccion,
            "tendencia": dict(tendencia_produccion),
            "registro": registro_produccion
        },
        "entrega": {
            "pendientes": ent_pendientes,
            "en_curso": ent_curso,
            "completadas": ent_completadas,
            "promedio_minutos": promedio_entrega,
            "eficiencia": eficiencia_entrega,
            "tendencia": dict(tendencia_entregas),
            "registro": registro_entregas
        },
        "despachos": {
            "en_proceso": ov_proceso,
            "listas": ov_lista,
            "despachadas": ov_despachada,
            "promedio_dias": promedio_dias,
            "eficiencia": eficiencia_despacho,
            "tendencia": dict(tendencia_despacho),
            "registro": registro_despachos
        },
        "consolidado_diario": consolidado_ordenado
    }
# ==========================================================
# ENDPOINT DASHBOARD (USA LA FUNCION)
# ==========================================================

@app.get("/api/dashboard_direccion")
def dashboard_direccion(
    request: Request,
    desde: str = Query(None),
    hasta: str = Query(None)
):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:

        # ======================================================
        # RANGO DE FECHAS
        # ======================================================
        if desde and hasta:
            try:
                fecha_desde = datetime.strptime(desde, "%Y-%m-%d")
                fecha_hasta = datetime.strptime(hasta, "%Y-%m-%d")
                fecha_hasta = fecha_hasta.replace(hour=23, minute=59, second=59)
            except ValueError:
                raise HTTPException(
                    status_code=400,
                    detail="Formato de fecha inválido. Use YYYY-MM-DD"
                )
        else:
            now = datetime.now()
            fecha_desde = datetime(now.year, now.month, 1)
            fecha_hasta = now

        # ======================================================
        # OBTENER MÉTRICAS
        # ======================================================
        data = obtener_metricas_dashboard(
            db,
            fecha_desde,
            fecha_hasta
        )

        data["rango"] = {
            "desde": fecha_desde.strftime("%Y-%m-%d"),
            "hasta": fecha_hasta.strftime("%Y-%m-%d")
        }

        return data

    finally:
        db.close()
# ==========================================================
# EXPORTAR DASHBOARD - PRODUCCIÓN (EXCEL)
# ==========================================================

from openpyxl import Workbook
from fastapi.responses import FileResponse
from fastapi import HTTPException
import os


@app.get("/dashboard/export/produccion")
def export_dashboard_produccion(
    request: Request,
    desde: str = Query(None),
    hasta: str = Query(None)
):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:

        # ======================================================
        # VALIDAR RANGO FECHAS
        # ======================================================
        if desde and hasta:
            try:
                fecha_desde = datetime.strptime(desde, "%Y-%m-%d")
                fecha_hasta = datetime.strptime(hasta, "%Y-%m-%d")
                fecha_hasta = fecha_hasta.replace(
                    hour=23,
                    minute=59,
                    second=59
                )
            except ValueError:
                raise HTTPException(
                    status_code=400,
                    detail="Formato de fecha inválido. Use YYYY-MM-DD"
                )
        else:
            now = datetime.now()
            fecha_desde = datetime(now.year, now.month, 1)
            fecha_hasta = now

        # ======================================================
        # OBTENER MÉTRICAS
        # ======================================================
        data = obtener_metricas_dashboard(
            db,
            fecha_desde,
            fecha_hasta
        )

        # ======================================================
        # CREAR EXCEL
        # ======================================================
        wb = Workbook()
        ws = wb.active
        ws.title = "Producción SGI"

        ws.append([
            "Pedido",
            "Cliente",
            "Total Piezas",
            "Escaneadas",
            "% Cumplimiento",
            "Última Fecha Escaneo"
        ])

        for r in data["produccion"].get("registro", []):
            ws.append([
                r["pedido"],
                r["cliente"],
                r["total_piezas"],
                r["escaneadas"],
                r["porcentaje"],
                r["ultima_fecha"]
            ])

        # ======================================================
        # GUARDAR ARCHIVO
        # ======================================================
        os.makedirs("reportes", exist_ok=True)

        timestamp = int(datetime.now().timestamp())

        filename = (
            f"REPORTE_PRODUCCION_"
            f"{fecha_desde.strftime('%Y%m%d')}_"
            f"{fecha_hasta.strftime('%Y%m%d')}_"
            f"{timestamp}.xlsx"
        )

        filepath = os.path.join("reportes", filename)

        wb.save(filepath)

        return FileResponse(
            filepath,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    finally:
        db.close()
# ==========================================================
# EXPORTAR DASHBOARD - ENTREGA CEDI (EXCEL)
# ==========================================================

@app.get("/dashboard/export/entrega")
def export_dashboard_entrega(
    request: Request,
    desde: str = Query(None),
    hasta: str = Query(None)
):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:

        # ======================================================
        # VALIDAR RANGO FECHAS
        # ======================================================
        if desde and hasta:
            try:
                fecha_desde = datetime.strptime(desde, "%Y-%m-%d")
                fecha_hasta = datetime.strptime(hasta, "%Y-%m-%d")
                fecha_hasta = fecha_hasta.replace(
                    hour=23,
                    minute=59,
                    second=59
                )
            except ValueError:
                raise HTTPException(
                    status_code=400,
                    detail="Formato de fecha inválido. Use YYYY-MM-DD"
                )
        else:
            now = datetime.now()
            fecha_desde = datetime(now.year, now.month, 1)
            fecha_hasta = now

        # ======================================================
        # OBTENER MÉTRICAS
        # ======================================================
        data = obtener_metricas_dashboard(
            db,
            fecha_desde,
            fecha_hasta
        )

        # ======================================================
        # CREAR EXCEL
        # ======================================================
        wb = Workbook()
        ws = wb.active
        ws.title = "Entrega CEDI"

        ws.append([
            "Pedido ID",
            "Responsable",
            "Inicio",
            "Fin",
            "Minutos",
            "Cumple (<=120)"
        ])

        for r in data["entrega"].get("registro", []):
            ws.append([
                r["pedido_id"],
                r["responsable"],
                r["inicio"],
                r["fin"],
                r["minutos"],
                r["cumple"]
            ])

        # ======================================================
        # GUARDAR ARCHIVO
        # ======================================================
        os.makedirs("reportes", exist_ok=True)

        timestamp = int(datetime.now().timestamp())

        filename = (
            f"REPORTE_ENTREGA_"
            f"{fecha_desde.strftime('%Y%m%d')}_"
            f"{fecha_hasta.strftime('%Y%m%d')}_"
            f"{timestamp}.xlsx"
        )

        filepath = os.path.join("reportes", filename)

        wb.save(filepath)

        return FileResponse(
            filepath,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    finally:
        db.close()
# ==========================================================
# EXPORTAR DASHBOARD - DESPACHOS (EXCEL)
# ==========================================================

@app.get("/dashboard/export/despachos")
def export_dashboard_despachos(
    request: Request,
    desde: str = Query(None),
    hasta: str = Query(None)
):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:

        # ======================================================
        # VALIDAR RANGO FECHAS
        # ======================================================
        if desde and hasta:
            try:
                fecha_desde = datetime.strptime(desde, "%Y-%m-%d")
                fecha_hasta = datetime.strptime(hasta, "%Y-%m-%d")
                fecha_hasta = fecha_hasta.replace(
                    hour=23,
                    minute=59,
                    second=59
                )
            except ValueError:
                raise HTTPException(
                    status_code=400,
                    detail="Formato de fecha inválido. Use YYYY-MM-DD"
                )
        else:
            now = datetime.now()
            fecha_desde = datetime(now.year, now.month, 1)
            fecha_hasta = now

        # ======================================================
        # OBTENER MÉTRICAS
        # ======================================================
        data = obtener_metricas_dashboard(
            db,
            fecha_desde,
            fecha_hasta
        )

        # ======================================================
        # CREAR EXCEL
        # ======================================================
        wb = Workbook()
        ws = wb.active
        ws.title = "Despachos"

        ws.append([
            "OV",
            "Fecha Lista",
            "Fecha Despacho",
            "Días en CEDI",
            "Cumple (<=5 días)"
        ])

        for r in data["despachos"].get("registro", []):
            ws.append([
                r["ov"],
                r["fecha_lista"],
                r["fecha_despacho"],
                r["dias"],
                r["cumple"]
            ])

        # ======================================================
        # GUARDAR ARCHIVO
        # ======================================================
        os.makedirs("reportes", exist_ok=True)

        timestamp = int(datetime.now().timestamp())

        filename = (
            f"REPORTE_DESPACHOS_"
            f"{fecha_desde.strftime('%Y%m%d')}_"
            f"{fecha_hasta.strftime('%Y%m%d')}_"
            f"{timestamp}.xlsx"
        )

        filepath = os.path.join("reportes", filename)

        wb.save(filepath)

        return FileResponse(
            filepath,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    finally:
        db.close()
# ==========================================================
# EXPORTAR DASHBOARD - CONSOLIDADO (EXCEL)
# ==========================================================

@app.get("/dashboard/export/consolidado")
def export_dashboard_consolidado(
    request: Request,
    desde: str = Query(None),
    hasta: str = Query(None)
):

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:

        # ======================================================
        # VALIDAR RANGO FECHAS
        # ======================================================
        if desde and hasta:
            try:
                fecha_desde = datetime.strptime(desde, "%Y-%m-%d")
                fecha_hasta = datetime.strptime(hasta, "%Y-%m-%d")
                fecha_hasta = fecha_hasta.replace(
                    hour=23,
                    minute=59,
                    second=59
                )
            except ValueError:
                raise HTTPException(
                    status_code=400,
                    detail="Formato de fecha inválido. Use YYYY-MM-DD"
                )
        else:
            now = datetime.now()
            fecha_desde = datetime(now.year, now.month, 1)
            fecha_hasta = now

        # ======================================================
        # OBTENER MÉTRICAS
        # ======================================================
        data = obtener_metricas_dashboard(
            db,
            fecha_desde,
            fecha_hasta
        )

        # ======================================================
        # CREAR EXCEL
        # ======================================================
        wb = Workbook()
        ws = wb.active
        ws.title = "Consolidado Diario"

        ws.append([
            "Fecha",
            "Producción",
            "Entregas",
            "Despachos"
        ])

        for fecha, valores in data["consolidado_diario"].items():
            ws.append([
                fecha,
                valores["produccion"],
                valores["entregas"],
                valores["despachos"]
            ])

        # ======================================================
        # GUARDAR ARCHIVO
        # ======================================================
        os.makedirs("reportes", exist_ok=True)

        timestamp = int(datetime.now().timestamp())

        filename = (
            f"REPORTE_CONSOLIDADO_"
            f"{fecha_desde.strftime('%Y%m%d')}_"
            f"{fecha_hasta.strftime('%Y%m%d')}_"
            f"{timestamp}.xlsx"
        )

        filepath = os.path.join("reportes", filename)

        wb.save(filepath)

        return FileResponse(
            filepath,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    finally:
        db.close()
# ==========================================================
# CALCULAR PERIODO ANTERIOR (COMPARATIVO MENSUAL REAL)
# ==========================================================

from datetime import datetime
from calendar import monthrange


def calcular_periodo_anterior(desde_dt, hasta_dt):
    """
    Retorna el rango completo del mes anterior
    basado en la fecha 'desde_dt'.
    """

    if not isinstance(desde_dt, datetime):
        raise ValueError("desde_dt debe ser datetime")

    mes_actual = desde_dt.month
    año_actual = desde_dt.year

    # Determinar mes anterior
    if mes_actual == 1:
        mes_anterior = 12
        año_anterior = año_actual - 1
    else:
        mes_anterior = mes_actual - 1
        año_anterior = año_actual

    # Último día del mes anterior
    ultimo_dia_anterior = monthrange(año_anterior, mes_anterior)[1]

    nuevo_desde = datetime(año_anterior, mes_anterior, 1)
    nuevo_hasta = datetime(
        año_anterior,
        mes_anterior,
        ultimo_dia_anterior,
        23,
        59,
        59
    )

    return nuevo_desde, nuevo_hasta


# ==========================================================
# SEGURIDAD GERENCIAL
# ==========================================================

PIN_GERENCIAL = "1308"
# ==========================================================
# PDF INFORME GENERAL SGI - VERSION CORPORATIVA FINAL
# ==========================================================

from fastapi import HTTPException, Query
from fastapi.responses import FileResponse
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import pagesizes
from reportlab.lib.colors import HexColor
from calendar import monthrange
from datetime import datetime
import os


@app.get("/dashboard/pdf-completo")
def generar_pdf_completo(
    request: Request,
    pin: str = Query(...),
    desde: str = Query(None),
    hasta: str = Query(None)
):

    # 🔐 PIN GERENCIAL
    if pin != PIN_GERENCIAL:
        raise HTTPException(status_code=403, detail="Acceso no autorizado")

    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:

        # ======================================================
        # RANGO FECHAS
        # ======================================================

        if desde and hasta:
            try:
                desde_dt = datetime.strptime(desde, "%Y-%m-%d")
                hasta_dt = datetime.strptime(hasta, "%Y-%m-%d")
                hasta_dt = hasta_dt.replace(hour=23, minute=59, second=59)
            except ValueError:
                raise HTTPException(
                    status_code=400,
                    detail="Formato de fecha inválido. Use YYYY-MM-DD"
                )
        else:
            now = datetime.now()
            ultimo_dia = monthrange(now.year, now.month)[1]
            desde_dt = datetime(now.year, now.month, 1)
            hasta_dt = datetime(now.year, now.month, ultimo_dia, 23, 59, 59)

        desde_ant, hasta_ant = calcular_periodo_anterior(desde_dt, hasta_dt)

        # ======================================================
        # MÉTRICAS
        # ======================================================

        data_actual = obtener_metricas_dashboard(db, desde_dt, hasta_dt)
        data_anterior = obtener_metricas_dashboard(db, desde_ant, hasta_ant)

        # ======================================================
        # CONFIGURACIÓN ARCHIVO
        # ======================================================

        carpeta = "reportes"
        os.makedirs(carpeta, exist_ok=True)

        timestamp = int(datetime.now().timestamp())

        filename = (
            f"INFORME_GENERAL_SGI_"
            f"{desde_dt.strftime('%Y%m%d')}_"
            f"{hasta_dt.strftime('%Y%m%d')}_"
            f"{timestamp}.pdf"
        )

        filepath = os.path.join(carpeta, filename)

        doc = SimpleDocTemplate(filepath, pagesize=pagesizes.A4)
        elements = []

        # ======================================================
        # ESTILOS
        # ======================================================

        style_normal = ParagraphStyle(
            name='NormalStyle',
            fontName='Helvetica',
            fontSize=11,
        )

        style_title = ParagraphStyle(
            name='TitleStyle',
            fontName='Helvetica-Bold',
            fontSize=18,
        )

        style_index = ParagraphStyle(
            name='IndexStyle',
            fontName='Helvetica-Bold',
            fontSize=26,
            textColor=HexColor("#ff7a00")
        )

        # ======================================================
        # DECORACIÓN CORPORATIVA
        # ======================================================

        def decoracion(canvas_obj, doc):

            canvas_obj.saveState()

            canvas_obj.setStrokeColor(HexColor("#ff7a00"))
            canvas_obj.setLineWidth(4)
            canvas_obj.rect(20, 20, 555, 802, stroke=1, fill=0)

            logo_path = "static/logo.png"
            if os.path.exists(logo_path):
                canvas_obj.drawImage(
                    logo_path,
                    420, 760,
                    width=120,
                    height=40,
                    preserveAspectRatio=True
                )

            canvas_obj.setFont("Helvetica-Bold", 70)
            canvas_obj.setFillColorRGB(1, 0.5, 0)
            try:
                canvas_obj.setFillAlpha(0.05)
            except:
                pass
            canvas_obj.drawCentredString(300, 420, "MADECENTRO")

            canvas_obj.restoreState()

        # ======================================================
        # PORTADA
        # ======================================================

        elements.append(Paragraph("INFORME GENERAL OPERATIVO SGI", style_title))
        elements.append(Spacer(1, 0.3 * inch))

        elements.append(Paragraph(
            f"Periodo Analizado: {desde_dt.strftime('%Y-%m-%d')} al {hasta_dt.strftime('%Y-%m-%d')}",
            style_normal
        ))

        elements.append(Spacer(1, 0.2 * inch))

        elements.append(Paragraph(
            f"ÍNDICE GENERAL SGI: {data_actual['indice_sgi']}%",
            style_index
        ))

        elements.append(Spacer(1, 0.5 * inch))

        # ======================================================
        # CONCLUSIÓN ESTRATÉGICA
        # ======================================================

        indice_actual = data_actual["indice_sgi"]
        indice_anterior = data_anterior["indice_sgi"]

        variacion = round(indice_actual - indice_anterior, 2)

        if variacion > 0:
            conclusion = f"El desempeño global mejora {variacion} puntos frente al mes anterior. "
        elif variacion < 0:
            conclusion = f"El desempeño global disminuye {abs(variacion)} puntos respecto al mes anterior. "
        else:
            conclusion = "El desempeño global se mantiene estable frente al mes anterior. "

        def estado_texto(valor):
            if valor >= 85:
                return "nivel óptimo"
            elif valor >= 65:
                return "nivel aceptable"
            else:
                return "nivel crítico"

        conclusion += f"Producción en {estado_texto(data_actual['produccion']['cumplimiento'])}. "
        conclusion += f"Entrega CEDI en {estado_texto(data_actual['entrega']['eficiencia'])}. "
        conclusion += f"Despachos en {estado_texto(data_actual['despachos']['eficiencia'])}."

        elements.append(Paragraph("CONCLUSIÓN ESTRATÉGICA", style_title))
        elements.append(Spacer(1, 0.2 * inch))
        elements.append(Paragraph(conclusion, style_normal))
        elements.append(Spacer(1, 0.5 * inch))

        # ======================================================
        # TABLAS (Producción, Entrega, Despachos)
        # ======================================================

        # PRODUCCIÓN
        elements.append(Paragraph("PRODUCCIÓN", style_title))
        elements.append(Spacer(1, 0.2 * inch))

        tabla_prod = Table([
            ["Métrica", "Actual", "Anterior"],
            ["Pendientes", data_actual["produccion"]["pendientes"], data_anterior["produccion"]["pendientes"]],
            ["En Proceso", data_actual["produccion"]["en_proceso"], data_anterior["produccion"]["en_proceso"]],
            ["Completadas", data_actual["produccion"]["completadas"], data_anterior["produccion"]["completadas"]],
            ["% Cumplimiento", data_actual["produccion"]["cumplimiento"], data_anterior["produccion"]["cumplimiento"]],
        ])

        tabla_prod.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.orange),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey)
        ]))

        elements.append(tabla_prod)
        elements.append(Spacer(1, 0.4 * inch))

        # ENTREGA
        elements.append(Paragraph("ENTREGA CEDI", style_title))
        elements.append(Spacer(1, 0.2 * inch))

        tabla_ent = Table([
            ["Métrica", "Actual", "Anterior"],
            ["Completadas", data_actual["entrega"]["completadas"], data_anterior["entrega"]["completadas"]],
            ["Promedio Min", data_actual["entrega"]["promedio_minutos"], data_anterior["entrega"]["promedio_minutos"]],
            ["% Eficiencia", data_actual["entrega"]["eficiencia"], data_anterior["entrega"]["eficiencia"]],
        ])

        tabla_ent.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.black),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey)
        ]))

        elements.append(tabla_ent)
        elements.append(Spacer(1, 0.4 * inch))

        # DESPACHOS
        elements.append(Paragraph("DESPACHOS", style_title))
        elements.append(Spacer(1, 0.2 * inch))

        tabla_des = Table([
            ["Métrica", "Actual", "Anterior"],
            ["Despachadas", data_actual["despachos"]["despachadas"], data_anterior["despachos"]["despachadas"]],
            ["Promedio Días", data_actual["despachos"]["promedio_dias"], data_anterior["despachos"]["promedio_dias"]],
            ["% Eficiencia", data_actual["despachos"]["eficiencia"], data_anterior["despachos"]["eficiencia"]],
        ])

        tabla_des.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), HexColor("#1f2937")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey)
        ]))

        elements.append(tabla_des)

        # ======================================================
        # GENERAR PDF
        # ======================================================

        doc.build(elements, onFirstPage=decoracion, onLaterPages=decoracion)

        return FileResponse(filepath, media_type='application/pdf', filename=filename)

    finally:
        db.close()
# ==========================================================
# ETIQUETAS PDF ZEBRA (SIN MARCO / TEXTO AMPLIO + MEDIDA + SUBPAQUETE)
# ==========================================================

from reportlab.graphics.barcode import code128
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from fastapi.responses import FileResponse
from fastapi import HTTPException
from datetime import datetime
import os


LABEL_WIDTH = 52 * mm
LABEL_HEIGHT = 70 * mm

PAGE_WIDTH = LABEL_WIDTH
PAGE_HEIGHT = LABEL_HEIGHT * 2   # DOS ETIQUETAS EN UNA HOJA

# ==========================================================
# OFFSET REAL (AJUSTE FINO)
# ==========================================================
OFFSET_X = -4.0 * mm
OFFSET_Y = -2.0 * mm


def generar_pdf_etiquetas(
    request: Request,
    pedido_id: int,
    paquete: str = None,
    codigo_unico: str = None
):
    planta = request.session.get("planta_codigo")

    if not planta:
        raise HTTPException(status_code=401, detail="Sesión no válida")

    db = get_db(planta)

    try:
        pedido = db.query(Pedido).filter(Pedido.id == pedido_id).first()
        if not pedido:
            raise HTTPException(status_code=404, detail="Pedido no encontrado")

        piezas_query = db.query(Pieza).filter(Pieza.pedido_id == pedido_id)

        # =========================
        # FILTROS
        # =========================
        if codigo_unico:
            pieza = piezas_query.filter(Pieza.codigo_unico == codigo_unico).first()
            if not pieza:
                raise HTTPException(status_code=404, detail="Código no encontrado en este pedido")
            piezas = [pieza]

        elif paquete:
            todas = piezas_query.order_by(Pieza.id.asc()).all()
            piezas = [
                p for p in todas
                if normalizar_paquete(p.paquete) == str(paquete).strip()
            ]

        else:
            piezas = piezas_query.order_by(Pieza.id.asc()).all()

        if not piezas:
            raise HTTPException(status_code=404, detail="No hay piezas para imprimir etiquetas")

        piezas.sort(
            key=lambda p: (
                key_paquete_num(normalizar_paquete(p.paquete)),
                p.codigo_unico
            )
        )

        # =========================
        # SUB-PAQUETE
        # =========================
        conteo_por_paquete = {}
        for p in piezas:
            pack = normalizar_paquete(p.paquete)
            conteo_por_paquete[pack] = conteo_por_paquete.get(pack, 0) + 1

        indice_por_paquete = {}
        for p in piezas:
            pack = normalizar_paquete(p.paquete)
            indice_por_paquete[pack] = indice_por_paquete.get(pack, 0) + 1
            p._sub_indice = indice_por_paquete[pack]
            p._sub_total = conteo_por_paquete[pack]

        # =========================
        # CREAR PDF
        # =========================
        os.makedirs("etiquetas_pdf", exist_ok=True)

        nombre_archivo = f"ETIQUETAS_{pedido.numero_pedido}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        ruta = os.path.join("etiquetas_pdf", nombre_archivo)

        c = canvas.Canvas(ruta, pagesize=(PAGE_WIDTH, PAGE_HEIGHT))

        def draw_center(text, font, size, y, y_offset):
            c.setFont(font, size)
            w = c.stringWidth(text, font, size)
            c.drawString(
                ((LABEL_WIDTH - w) / 2) + OFFSET_X,
                (y_offset + y) + OFFSET_Y,
                text
            )

        # =========================
        # DIBUJAR ETIQUETAS
        # =========================
        for i, pieza in enumerate(piezas):

            y_offset = LABEL_HEIGHT if i % 2 == 0 else 0

            pedido_txt = safe_text(pedido.numero_pedido, 25)
            paquete_txt = normalizar_paquete(pieza.paquete)
            sub = f"{pieza._sub_indice}/{pieza._sub_total}"

            largo = getattr(pieza, "base", None)
            ancho = getattr(pieza, "altura", None)

            try:
                largo = int(float(str(largo))) if largo else 0
            except:
                largo = 0

            try:
                ancho = int(float(str(ancho))) if ancho else 0
            except:
                ancho = 0

            medida_txt = (
                f"MEDIDA: {largo} x {ancho}"
                if largo > 0 and ancho > 0
                else "MEDIDA: SIN MEDIDA"
            )

            y_top = LABEL_HEIGHT - 8 * mm

            draw_center(f"PEDIDO: {pedido_txt}", "Helvetica-Bold", 10, y_top, y_offset)
            draw_center(f"PAQUETE: {paquete_txt} ({sub})", "Helvetica-Bold", 10, y_top - 7 * mm, y_offset)
            draw_center(medida_txt, "Helvetica-Bold", 8.5, y_top - 13 * mm, y_offset)

            # =========================
            # ✅ EAN-13 USANDO EXACTAMENTE codigo_base_12 DE BD
            # =========================
            codigo_base = (pieza.codigo_base_12 or "").strip()

            if len(codigo_base) != 12:
                raise Exception(f"codigo_base_12 inválido: {codigo_base}")

            barcode = eanbc.Ean13BarcodeWidget(codigo_base)
            barcode.barHeight = 26 * mm
            barcode.barWidth = 0.34 * mm
            barcode.humanReadable = True

            bounds = barcode.getBounds()
            barcode_width = bounds[2] - bounds[0]
            barcode_height = bounds[3] - bounds[1]

            d = Drawing(barcode_width, barcode_height)
            d.add(barcode)

            x_barcode = ((LABEL_WIDTH - barcode_width) / 2) + OFFSET_X
            y_barcode = y_offset + (18 * mm) + OFFSET_Y

            renderPDF.draw(d, c, x_barcode, y_barcode)

            if i % 2 == 1:
                c.showPage()

        c.save()
        return ruta

    finally:
        db.close()
# ==========================================================
# ENDPOINTS ETIQUETAS PDF
# ==========================================================

@app.get("/etiquetas_pdf/todas/{pedido_id}")
def etiquetas_todas(request: Request, pedido_id: int):
    ruta = generar_pdf_etiquetas(
        request=request,
        pedido_id=pedido_id
    )
    return FileResponse(
        ruta,
        media_type="application/pdf",
        filename=os.path.basename(ruta)
    )


@app.get("/etiquetas_pdf/paquete/{pedido_id}/{paquete}")
def etiquetas_paquete(request: Request, pedido_id: int, paquete: str):
    ruta = generar_pdf_etiquetas(
        request=request,
        pedido_id=pedido_id,
        paquete=paquete
    )
    return FileResponse(
        ruta,
        media_type="application/pdf",
        filename=os.path.basename(ruta)
    )


@app.get("/etiquetas_pdf/individual/{pedido_id}/{codigo_unico}")
def etiquetas_individual(request: Request, pedido_id: int, codigo_unico: str):
    ruta = generar_pdf_etiquetas(
        request=request,
        pedido_id=pedido_id,
        codigo_unico=codigo_unico
    )
    return FileResponse(
        ruta,
        media_type="application/pdf",
        filename=os.path.basename(ruta)
    )

# ==========================================================
# VISTA DASHBOARD DIRECCIÓN
# ==========================================================

@app.get("/dashboard", response_class=HTMLResponse)
def vista_dashboard(request: Request):

    if not request.session.get("planta_codigo"):
        return RedirectResponse(url="/login", status_code=302)

    return templates.TemplateResponse(
        "dashboard.html",
        {"request": request}
    )


# ==========================================================
# CREAR TABLAS EN TODAS LAS PLANTAS (SOLO DESARROLLO)
# ==========================================================

if os.getenv("ENV") != "PRODUCCION":
    for engine in engines.values():
        Base.metadata.create_all(bind=engine)